import os, json, datetime
import pandas as pd
import pyxlsb
import openpyxl
import requests

FP = "data/budget.xlsb"
CATEGORIES_SHEET = "Categories"
FP_JRF = "data/jrf_budget.xlsx"

# ── A/P - CNS (live, from Airtable) ───────────────────────────────────────────
# Overseas COGS purchase-order payables tracked in the "Ledger 💰" Airtable
# base -- real vendor debt that hasn't hit the Cash Flow - Tracker's cash-out
# actuals yet (it only shows up there once actually paid), so it's the piece
# missing from a pure spreadsheet-based cash projection.
#
# KNOWN GAP (2026-08-24): this table only covers CNS vendors. It does not
# include key blank-goods vendors (S&S Activewear, SanMar, As Colour, etc.)
# -- real COGS - Brand cash-out, confirmed showing up week after week in the
# Xero GL exports -- so "live A/P" here understates true outstanding
# COGS - Brand payables, not overstates. Flagged by Jeremy: COGS needs a
# better forecast overall, since payment *timing* (not just PO/invoice
# amount) is what actually drives near-term cash out. Not solved yet --
# needs either a second A/P source for blank-goods vendors or a
# timing-aware model, not just a bigger balance to swap in.
AIRTABLE_API = "https://api.airtable.com/v0"
AIRTABLE_LEDGER_BASE = "appW8jAfERj3iBzqt"  # Ledger 💰
AP_CNS_TABLE = "tblC9gQD9rrRbVtno"
# "Payment Schedule ⌚" -- Jeremy's own view, scoped to A/P - CNS with a
# confirmed due date. Querying it directly (via the `view` param) instead of
# reconstructing an equivalent filterByFormula, since the two silently
# diverged: the old NOT({Paid in Full}) + due-date-not-blank filter also
# pulled in unpaid records that only look due-dated once you dig in (or
# don't really belong on a payment schedule at all) -- ~$1.26M sitting in
# records with no confirmed due date, entirely excluded from both old and
# new totals, but no longer silently conflated with the confirmed-due set.
AP_CNS_PAYMENT_SCHEDULE_VIEW = "viwPmOdLMRteANj8v"

def fetch_ap_cns():
    """Unpaid A/P - CNS balances with a confirmed due date (the "Payment
    Schedule" view), bucketed into overdue vs. due by year-end. Returns None
    (rather than raising) if AIRTABLE_API_KEY isn't configured or the
    request fails, so a local/offline build just omits the Paydown
    Feasibility section instead of breaking."""
    token = os.environ.get('AIRTABLE_API_KEY')
    if not token:
        return None
    url = f"{AIRTABLE_API}/{AIRTABLE_LEDGER_BASE}/{AP_CNS_TABLE}"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "view": AP_CNS_PAYMENT_SCHEDULE_VIEW,
        "fields[]": ["⚠ Due Date", "Remaining Due"],
        "pageSize": 100,
    }
    records = []
    try:
        offset = None
        while True:
            p = dict(params)
            if offset:
                p["offset"] = offset
            r = requests.get(url, headers=headers, params=p, timeout=30)
            r.raise_for_status()
            data = r.json()
            records.extend(data.get("records", []))
            offset = data.get("offset")
            if not offset:
                break
    except requests.RequestException as e:
        print("A/P - CNS fetch failed, skipping paydown feasibility:", e)
        return None

    today = datetime.date.today()
    year_end = datetime.date(today.year, 12, 31)
    overdue = upcoming = beyond_year = 0.0
    for rec in records:
        f = rec.get("fields", {})
        remaining = f.get("Remaining Due") or 0
        due_str = f.get("⚠ Due Date")
        if not due_str or not remaining:
            continue
        due = datetime.date.fromisoformat(due_str)
        if due < today:
            overdue += remaining
        elif due <= year_end:
            upcoming += remaining
        else:
            beyond_year += remaining
    return dict(overdue=overdue, upcoming=upcoming, beyond_year=beyond_year,
                total=overdue + upcoming, as_of=today.isoformat())

# ── A/P - Key Vendors (live, from Airtable) ───────────────────────────────────
# Domestic blank-goods vendors (S&S Activewear, SanMar, AS Colour, etc.),
# purchased through Jetty INK, covering both Brand's own product and INK's
# private-label/screen-printing work -- the domestic counterpart to A/P-CNS's
# overseas Cut & Sew production. Two live sources, both in Airtable:
#  - "Ledger" base's BANK table: a hand-entered weekly aging snapshot
#    (Current/1-30/31-60/61-90/90+), split BRAND vs INK, back to Week 1 of
#    2026 -- the long-run trend, but not broken out per vendor.
#  - "Brand Production" base's WHSL PMT Tracker + Vendor List: live, per-PO
#    detail (vendor, due date, INK $ / Brand $, paid status) plus each
#    vendor's stated payment terms -- real-time per-vendor exposure and
#    aging, computed here from today's date rather than Airtable's own live
#    "Days Past Due" formula, so a build is reproducible against the data as
#    of when it ran, not whenever someone happens to re-open the base.
AIRTABLE_BANK_TABLE = "tbl4U2p2USUU9Qst6"
AIRTABLE_BRAND_PROD_BASE = "applvN4vFAPqnDwK2"  # Brand Production
VENDOR_LIST_TABLE = "tblOk4r79oc9Hsv4L"
WHSL_PMT_TABLE = "tblIIvVAkqGOSeKrR"

# How overdue a Key Vendor's balance has to be before it's a real must-pay.
# Jeremy pays most vendors once they're 31+ days late, but has a standing
# arrangement with S&S Activewear where only 90+ has to go out. Drives both
# the Summary panel's "not yet due" provision and which aging columns are
# flagged red in the Open POs by Vendor table -- one rule, one place.
SS_VENDOR_NAME = 'S&S Activewear'
KV_MUST_PAY_BUCKETS = ('d31_60', 'd61_90', 'd90_plus')
KV_MUST_PAY_BUCKETS_SS = ('d90_plus',)

def kv_must_pay_buckets(vendor):
    """Which aging buckets this vendor actually has to pay right now."""
    return KV_MUST_PAY_BUCKETS_SS if vendor == SS_VENDOR_NAME else KV_MUST_PAY_BUCKETS

def _airtable_get_all(base, table, token, params):
    """Paginate through every record for a table/filter, return the raw list."""
    url = f"{AIRTABLE_API}/{base}/{table}"
    headers = {"Authorization": f"Bearer {token}"}
    records = []
    offset = None
    while True:
        p = dict(params)
        if offset:
            p["offset"] = offset
        r = requests.get(url, headers=headers, params=p, timeout=30)
        r.raise_for_status()
        data = r.json()
        records.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            break
    return records

def fetch_ap_key_vendors_weekly():
    """Weekly A/P - Key Vendors aging (BRAND + INK, each broken out), from the
    Ledger base's BANK table -- the same Current/1-30/31-60/61-90/90+
    structure as the AR tab, just for domestic blank-goods payables instead
    of receivables. Returns None on any failure so the build degrades
    gracefully, same convention as fetch_ap_cns()."""
    token = os.environ.get('AIRTABLE_API_KEY')
    if not token:
        return None
    params = {
        "filterByFormula": "AND(OR({Category}='BRAND',{Category}='INK'), {Year}='2026', {Total Aged A/P}>0)",
        "fields[]": ["Week", "Category", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days"],
        "pageSize": 100,
    }
    try:
        records = _airtable_get_all(AIRTABLE_LEDGER_BASE, AIRTABLE_BANK_TABLE, token, params)
    except requests.RequestException as e:
        print("A/P - Key Vendors weekly fetch failed, skipping:", e)
        return None

    def empty_bucket():
        return dict(current=0.0, d1_30=0.0, d31_60=0.0, d61_90=0.0, d90_plus=0.0, total=0.0)

    by_week = {}
    for rec in records:
        f = rec.get("fields", {})
        wk, cat = f.get("Week"), f.get("Category")
        if wk is None or cat not in ("BRAND", "INK"):
            continue
        entry = by_week.setdefault(int(wk), dict(wk=int(wk), brand=None, ink=None))
        bucket = dict(current=f.get("Current", 0) or 0, d1_30=f.get("1-30 Days", 0) or 0,
                      d31_60=f.get("31-60 Days", 0) or 0, d61_90=f.get("61-90 Days", 0) or 0,
                      d90_plus=f.get("90+ Days", 0) or 0)
        bucket["total"] = sum(bucket.values())
        entry["brand" if cat == "BRAND" else "ink"] = bucket
    weekly = [by_week[wk] for wk in sorted(by_week)]
    for w in weekly:
        w["brand"] = w["brand"] or empty_bucket()
        w["ink"]   = w["ink"]   or empty_bucket()
        w["total"] = w["brand"]["total"] + w["ink"]["total"]
    return weekly

def fetch_ap_key_vendors_terms():
    """Vendor name -> stated payment terms (e.g. 'NET30', 'NET60', 'Credit
    Card'), from Brand Production's Vendor List. Returns {} on failure."""
    token = os.environ.get('AIRTABLE_API_KEY')
    if not token:
        return {}
    params = {"fields[]": ["Name", "Terms"], "pageSize": 100}
    try:
        records = _airtable_get_all(AIRTABLE_BRAND_PROD_BASE, VENDOR_LIST_TABLE, token, params)
    except requests.RequestException as e:
        print("Vendor List fetch failed, skipping terms:", e)
        return {}
    terms = {}
    for rec in records:
        f = rec.get("fields", {})
        name = f.get("Name")
        t = f.get("Terms") or []
        if name:
            terms[name] = t[0] if t else None
    return terms

def fetch_ap_key_vendors_open():
    """Every currently-unpaid PO in WHSL PMT Tracker, grouped by vendor and
    bucketed into Current/1-30/31-60/61-90/90+ from each PO's Due Date vs.
    today -- live per-vendor aging, which the weekly BANK snapshot doesn't
    give. Each PO's dollars are split into its INK vs. Brand-side share (the
    table's own INK / Total BRAND A/P columns), so the totals returned here
    can feed each side's own trend line without double-counting or guessing
    a split. Returns None on failure."""
    token = os.environ.get('AIRTABLE_API_KEY')
    if not token:
        return None
    params = {
        "filterByFormula": "NOT({Paid})",
        "fields[]": ["Vendor", "Due Date", "INK", "Total BRAND A/P"],
        "pageSize": 100,
    }
    try:
        records = _airtable_get_all(AIRTABLE_BRAND_PROD_BASE, WHSL_PMT_TABLE, token, params)
    except requests.RequestException as e:
        print("WHSL PMT Tracker fetch failed, skipping open A/P detail:", e)
        return None

    today = datetime.date.today()
    by_vendor = {}
    for rec in records:
        f = rec.get("fields", {})
        vendor_list = f.get("Vendor") or []
        vendor = vendor_list[0] if vendor_list else "(unknown)"
        ink_amt = f.get("INK") or 0
        brand_amt = f.get("Total BRAND A/P") or 0
        amt = ink_amt + brand_amt
        due_str = f.get("Due Date")
        if not amt or not due_str:
            continue
        due = datetime.date.fromisoformat(due_str)
        days_past = (today - due).days
        v = by_vendor.setdefault(vendor, dict(vendor=vendor, current=0.0, d1_30=0.0, d31_60=0.0,
                                               d61_90=0.0, d90_plus=0.0, ink_total=0.0, brand_total=0.0,
                                               count=0, worst_days=days_past))
        bucket_key = ('current' if days_past <= 0 else 'd1_30' if days_past <= 30 else
                      'd31_60' if days_past <= 60 else 'd61_90' if days_past <= 90 else 'd90_plus')
        v[bucket_key] += amt
        v["ink_total"] += ink_amt
        v["brand_total"] += brand_amt
        v["count"] += 1
        v["worst_days"] = max(v["worst_days"], days_past)
    for v in by_vendor.values():
        v["total"] = v["current"] + v["d1_30"] + v["d31_60"] + v["d61_90"] + v["d90_plus"]
    vendors = sorted(by_vendor.values(), key=lambda v: -v["total"])
    return dict(as_of=today.isoformat(), vendors=vendors,
                total=sum(v["total"] for v in vendors),
                ink_total=sum(v["ink_total"] for v in vendors),
                brand_total=sum(v["brand_total"] for v in vendors))

# ── Formatters ──────────────────────────────────────────────────────────────

def html_escape(s):
    return (s.replace('&', '&amp;').replace('"', '&quot;')
             .replace('<', '&lt;').replace('>', '&gt;'))

def fk(v):
    if v == 0: return "$0"
    a = abs(v); s = "minus" if v < 0 else ""
    if a >= 1_000_000: return ("" if v>0 else "−") + "$" + format(a/1_000_000, ".2f") + "M"
    if a >= 1_000:     return ("" if v>0 else "−") + "$" + format(a/1_000,     ".1f") + "K"
    return ("" if v>0 else "−") + "$" + f"{a:,.0f}"

def vk(v):
    if v == 0: return "—"
    sign = "+" if v > 0 else "−"; a = abs(v)
    if a >= 1_000_000: return sign + "$" + format(a/1_000_000, ".2f") + "M"
    if a >= 1_000:     return sign + "$" + format(a/1_000,     ".1f") + "K"
    return sign + "$" + f"{a:,.0f}"
TAB_JS = '''
document.querySelectorAll(".tab").forEach(function(t){
  t.addEventListener("click", function(){
    document.querySelectorAll(".tab").forEach(function(x){x.classList.remove("active");});
    document.querySelectorAll(".tab-panel").forEach(function(x){x.classList.remove("active");});
    t.classList.add("active");
    var panel = document.getElementById("panel-" + t.dataset.panel);
    if (panel) panel.classList.add("active");
    if (window.Chart && Chart.instances) {
      Object.values(Chart.instances).forEach(function(c){ c.resize(); });
    }
  });
});
'''

# A single reusable full-screen modal: any chart registered into
# window.__charts[chartId] (see build_summary_trend_charts_js) can be
# blown up full-screen with a PNG download, via the card's Expand button
# (see rc_trend_chart).
CHART_MODAL_HTML = '''
<div class="chart-modal-overlay" id="chart-modal">
  <div class="chart-modal-content">
    <div class="chart-modal-header">
      <div class="chart-modal-title" id="chart-modal-title"></div>
      <div class="chart-modal-actions">
        <button onclick="downloadChartModal()">⬇ Download PNG</button>
        <button class="chart-modal-close" onclick="closeChartModal()">✕</button>
      </div>
    </div>
    <div class="chart-modal-canvas-wrap"><canvas id="chart-modal-canvas"></canvas></div>
  </div>
</div>
'''

CHART_MODAL_JS = '''
window.__charts = window.__charts || {};
window.__modalChart = null;
function openChartModal(chartId, title){
  var src = window.__charts[chartId];
  var data = window.__chartData[chartId];
  var optsFn = window.__chartOptsFn[chartId];
  if(!src || !data) return;
  document.getElementById("chart-modal-title").textContent = title;
  document.getElementById("chart-modal").classList.add("open");
  if(window.__modalChart){ window.__modalChart.destroy(); window.__modalChart = null; }
  var ctx = document.getElementById("chart-modal-canvas");
  window.__modalChart = new Chart(ctx, {
    type: src.config.type,
    data: data,
    options: optsFn ? optsFn() : {responsive:true, maintainAspectRatio:false}
  });
  window.__modalTitle = title;
}
function closeChartModal(){
  document.getElementById("chart-modal").classList.remove("open");
  if(window.__modalChart){ window.__modalChart.destroy(); window.__modalChart = null; }
}
function downloadChartModal(){
  if(!window.__modalChart) return;
  // toBase64Image()'s canvas is transparent -- fine on a white page, but
  // most image viewers (and dark-mode ones especially) render transparency
  // as black, making the light gridlines/text unreadable. Composite a white
  // fill behind the existing pixels ("destination-over" draws under, not
  // over), grab the URL, then redraw the chart to undo it on-screen.
  var canvas = window.__modalChart.canvas;
  var ctx = canvas.getContext("2d");
  ctx.save();
  ctx.globalCompositeOperation = "destination-over";
  ctx.fillStyle = "#ffffff";
  ctx.fillRect(0, 0, canvas.width, canvas.height);
  var url = canvas.toDataURL("image/png", 1);
  ctx.restore();
  window.__modalChart.draw();
  var link = document.createElement("a");
  link.download = (window.__modalTitle || "chart").replace(/[^a-z0-9]+/gi,"_") + ".png";
  link.href = url;
  link.click();
}
document.getElementById("chart-modal").addEventListener("click", function(e){
  if(e.target.id === "chart-modal") closeChartModal();
});
document.addEventListener("click", function(e){
  var btn = e.target.closest(".rc-expand-btn");
  if(btn) openChartModal(btn.dataset.chartId, btn.dataset.chartTitle);
});
document.addEventListener("keydown", function(e){
  if(e.key === "Escape") closeChartModal();
});
'''

def js_arr(lst):
    return '[' + ','.join(str(v) for v in lst) + ']'

def to_xy(d, est=None):
    est = est or set()
    return '[' + ','.join('{x:"Wk ' + str(k) + '",y:' + str(v) + ',e:' + ('true' if k in est else 'false') + '}'
                           for k,v in sorted(d.items())) + ']'

def st_xy(d, est=None):
    est = est or set()
    return '[' + ','.join('{x:"Wk ' + str(k) + '",y:' + str(v) + ',e:' + ('true' if k in est else 'false') + '}'
                           for k,v in sorted(d.items()) if v) + ']'

def trend_xy(arr):
    """arr is a 52-length list (index 0 = week 1), None where there's no data
    for that week -- those weeks are omitted entirely so Chart.js draws a
    gap/skip rather than a fake drop to zero."""
    return '[' + ','.join('{x:"Wk ' + str(i+1) + '",y:' + repr(round(float(v), 2)) + '}'
                           for i, v in enumerate(arr) if v is not None) + ']'

# ── Seasonality / trend ───────────────────────────────────────────────────────
# The dashboard's default "Ann. Proj." (used everywhere for variance reporting)
# is deliberately naive: actual-to-date + whatever's still budgeted for the
# rest of the year, so Ann. Var always equals YTD Var. That's fine for
# tracking plan adherence but it's blind to how a channel is actually
# pacing — a channel already 10% ahead of plan gets projected as if it will
# suddenly land exactly on budget for every remaining week.
#
# Trend Proj. instead asks "given this channel's normal shape for the year,
# and how much of that shape's ground we've covered so far, where does our
# actual pace put us by December?" — ytd_actual / (expected % of the year
# already behind us). That naturally builds in trend (over/under-pace shows
# up directly) *and* seasonality (each channel's "expected % so far" is its
# own curve), which is why one channel's early-year pace means something
# very different from another's.
#
# CHANNEL_SEASONALITY below is derived from real 2025 actuals (the "2025
# Actual" tab), not guessed — each channel's monthly weight is its share of
# that channel's full 2025 revenue, bucketed into the same 4-4-5 retail
# calendar the sheet itself uses for weeks (verified against the Inventory
# tab's Start/End columns).
#
# One wrinkle: Long Branch opened mid-2025 and its revenue was booked
# lumped into Flagship Store all year in the books (no separate ledger
# line), so "DTC - Flagship Store" in 2025 Actual is really Flagship+Long
# Branch combined. Rather than estimate a split, we pulled each store's
# real daily Net Sales straight from Airtable — the Jetty Hub base's
# "DTC Daily" table tracks both stores separately by day, with a Sales
# Channel of "Flagship Store" or "Long Branch" — and used those actual
# per-store weekly totals for each store's own curve. (The two stores'
# 2025 total, summed, comes in within 1% of the lumped ledger figure —
# a good cross-check that the two sources agree.)
#
# Weeks are bucketed into a standard 4-4-5 retail calendar (13 weeks/quarter,
# 52 total), matching the sheet's own week numbering exactly.
MONTH_WEEKS = [4,4,5, 4,4,5, 4,4,5, 4,4,5]  # Jan..Dec

CHANNEL_SEASONALITY = {
    # Jan,   Feb,   Mar,   Apr,   May,   Jun,   Jul,   Aug,   Sep,   Oct,   Nov,   Dec
    'DTC - Jettylife.com':       [.038,.033,.057,.049,.071,.078,.061,.061,.121,.081,.097,.251],
    'Wholesale Revenue':         [.016,.031,.187,.090,.067,.077,.141,.128,.129,.072,.042,.019],
    'Screen Printing Revenue':   [.029,.041,.139,.083,.078,.117,.103,.099,.156,.062,.081,.013],
    'DTC - Flagship Store':      [.021,.036,.054,.066,.072,.107,.092,.129,.088,.045,.088,.202],
    'DTC - Long Branch':         [.000,.000,.000,.000,.069,.212,.203,.205,.121,.053,.057,.080],
    'DTC - Mobile Store & Tent': [.015,.012,.065,.026,.052,.081,.098,.151,.070,.337,.045,.047],
    'JRF - Screen Printing':     [.144,.141,.035,.000,.000,.100,.058,.049,.221,.142,.092,.018],

    # NOTE: we tried adding aggregate COGS_LS_TOTAL / OPEX_TOTAL curves here,
    # derived from the 2025 Actual tab's Total Cost of Goods Sold / Total
    # Operating Expenses rows the same way as the channels above. Reverted --
    # unlike revenue, COGS/Labor/OpEx timing isn't driven by a repeatable
    # calendar pattern, it's driven by when bills get entered into the books.
    # 2025's shape only looked back-loaded because Fall/Winter COGS was
    # entered late that year; 2026's budget already has ~72% of its annual
    # COGS+Labor+Shipping total booked by week 31 (large lump-sum weeks, e.g.
    # $1.26M in week 30) rather than spread seasonally. Extrapolating off
    # 2025's entry-timing artifact produced a full-year COGS+Labor+Shipping
    # projection ~$2.5M over the actual trajectory. Revenue's curves above are
    # fine -- consumer buying seasonality genuinely repeats year to year.
}

def _weekly_weights(monthly):
    out = []
    for wk_count, m_wt in zip(MONTH_WEEKS, monthly):
        out.extend([m_wt / wk_count] * wk_count)
    return out  # 52 weights, summing to ~1.0

def seasonal_ann_proj(key, ytd_act, ann_plan, wk):
    """Trend + seasonality-adjusted full-year projection for one channel.
    Falls back to the plan itself if there's no curve or no ground covered
    yet, rather than divide-by-zero-ing into nonsense in week 1."""
    monthly = CHANNEL_SEASONALITY.get(key)
    if not monthly:
        return ann_plan
    cum = sum(_weekly_weights(monthly)[:wk])
    return (ytd_act / cum) if cum > 0 else ann_plan

# ── Data extraction ──────────────────────────────────────────────────────────

def read_categories():
    """Optional 'Categories' tab: <line item> | Category | Highlight | [Callout Label].
    Columns are matched by position, not header text, and a title/blank row above
    the header is tolerated — the header row is whichever row has "categ" in its
    second cell. The line item must match the sheet name used in the Actual/Budget
    tabs exactly. Returns {} if the tab doesn't exist yet, so the OpEx section
    falls back to the flat opex_map-driven table."""
    try:
        raw = pd.read_excel(FP, sheet_name=CATEGORIES_SHEET, engine='pyxlsb', header=None)
    except Exception:
        return {}
    header_row = None
    for i, row in raw.iterrows():
        if pd.notna(row.get(1)) and 'categ' in str(row[1]).strip().lower():
            header_row = i
            break
    if header_row is None:
        return {}
    cats = {}
    for _, row in raw.iloc[header_row + 1:].iterrows():
        name = row.get(0)
        if pd.isna(name) or not str(name).strip():
            continue
        category = row.get(1)
        highlight = pd.notna(row.get(2)) and str(row.get(2)).strip() != ''
        callout = row.get(3) if len(row) > 3 else None
        cats[str(name).strip()] = {
            'category': str(category).strip() if pd.notna(category) and str(category).strip() else 'Uncategorized',
            'highlight': highlight,
            'callout': str(callout).strip() if highlight and pd.notna(callout) and str(callout).strip() else None,
        }
    return cats

# ── Cash Flow tab rebuild (Sept 2026) ─────────────────────────────────────────
# Cash Flow - Weekly's layout changed when Jeremy added full 2025 weekly
# history alongside 2026: it's now Year-blocked (col 0 = Year, col 1 =
# Account, cols 2-53 = weeks 1-52, cols 54-60 = Q1..ANNUAL rollups), each
# year contributing 5 account rows. The old single-year cfw_row() lookup
# (still used by bank_position et al. until those sections are rebuilt too)
# just grabs the first row matching a label, which now silently picks up
# 2025's row before 2026's -- the root cause of "something is messed up
# now". This reader is self-contained and doesn't touch that old path.
def read_cash_flow_weekly_history(fp):
    """Returns {year: {'cash_in': [wk1..wk52], 'cash_out': [...],
    'cl_draw': [...], 'cl_paydown': [...], 'cl_balance': [...]}}, values
    None for weeks not yet entered. Returns {} if the sheet is missing or
    empty. 'cl_balance' (the sheet's own running Columbia CL balance row)
    is fully populated through week 52 -- it holds flat at the last real
    draw/paydown rather than going blank for future weeks, unlike the
    other four series here.
    (Bank balance history lives on the separate 'Bank Balance' sheet --
    see read_bank_balance_history -- the 'Total Starting Bank Balance'
    row on this sheet is currently unpopulated.)"""
    try:
        df = pd.read_excel(fp, sheet_name='Cash Flow - Weekly', engine='pyxlsb', header=None)
    except ValueError:
        return {}
    accounts = {'Total Cash IN': 'cash_in', 'Total Cash OUT': 'cash_out',
                'Columbia CL Draw': 'cl_draw', 'Columbia CL Paydown': 'cl_paydown',
                'Columbia CL Balance': 'cl_balance'}
    by_year = {}
    current_year = None
    for _, row in df.iterrows():
        yr, acct = row[0], row[1]
        if pd.notna(yr):
            try:
                current_year = int(yr)
            except (TypeError, ValueError):
                current_year = None
        key = accounts.get(str(acct).strip()) if pd.notna(acct) else None
        if current_year and key:
            weeks = [float(row[c]) if pd.notna(row[c]) else None for c in range(2, 54)]
            by_year.setdefault(current_year, {})[key] = weeks
    return by_year

def read_bank_balance_history(fp):
    """Returns {year: [wk1..wk52]} of week-ending total bank balance
    (Columbia + BOA + Ramp) from the 'Bank Balance' sheet's 'Total'
    column, keyed by that sheet's own 'Week' number (1-52 per year).
    Trailing zeros -- the sheet's own SUM formula returns 0, not blank,
    for weeks not yet reached -- are converted to None so a balance line
    stops rather than dropping to zero. Returns {} if the sheet is
    missing/unreadable."""
    try:
        df = pd.read_excel(fp, sheet_name='Bank Balance', engine='pyxlsb', header=None)
    except ValueError:
        return {}
    by_year = {}
    for _, row in df.iterrows():
        yr, wk, total = row[0], row[2], row[8]
        try:
            yr, wk = int(yr), int(wk)
        except (TypeError, ValueError):
            continue
        if not (1 <= wk <= 52):
            continue
        by_year.setdefault(yr, [None] * 52)[wk - 1] = float(total) if pd.notna(total) else None
    for weeks in by_year.values():
        for i in range(len(weeks) - 1, -1, -1):
            if weeks[i] == 0:
                weeks[i] = None
            else:
                break
    return by_year

def read_ar_history(fp):
    """Returns {year: {'brand_total': [wk1..wk52], 'ink_total': [...]}} of
    weekly Total A/R (the 'Total' column of each block, not the '0-90
    Days' subtotal -- includes 91+ Days), from the 'AR' sheet's Brand A/R
    and INK A/R blocks, keyed by that sheet's own 'Week' number. Trailing
    zeros (weeks not yet reached) are converted to None, same convention
    as read_bank_balance_history(). Returns {} if the sheet is missing/
    unreadable. Unlike Cash Flow/Bank Balance, this sheet only has a 2026
    block -- no prior-year comparison is possible from this source."""
    try:
        df = pd.read_excel(fp, sheet_name='AR', engine='pyxlsb', header=None)
    except ValueError:
        return {}
    by_year = {}
    for _, row in df.iterrows():
        yr, wk, brand_tot, ink_tot = row[0], row[2], row[9], row[16]
        try:
            yr, wk = int(yr), int(wk)
        except (TypeError, ValueError):
            continue
        if not (1 <= wk <= 52):
            continue
        entry = by_year.setdefault(yr, {'brand_total': [None] * 52, 'ink_total': [None] * 52})
        entry['brand_total'][wk - 1] = float(brand_tot) if pd.notna(brand_tot) else None
        entry['ink_total'][wk - 1] = float(ink_tot) if pd.notna(ink_tot) else None
    for entry in by_year.values():
        for key in ('brand_total', 'ink_total'):
            weeks = entry[key]
            for i in range(len(weeks) - 1, -1, -1):
                if weeks[i] == 0:
                    weeks[i] = None
                else:
                    break
    return by_year

def read_ap_vendor_total(fp, sheet_name):
    """Returns {year: [wk1..52]} of the TOTAL row from the 'AP - Key' /
    'AP - CNS' sheets -- Year/Type/Terms/Account header, one row per
    vendor, weekly columns 1-52, a 'TOTAL A/P - ...' row at the bottom.
    These are hand-entered cumulative snapshots (not backfilled week by
    week), so unlike the other read_* helpers here this returns the raw
    values as-is, zeros included -- callers should take the latest
    non-zero week as "the current balance" rather than assuming a
    contiguous run from week 1. Returns {} if the sheet is missing/
    unreadable or doesn't match the expected layout."""
    try:
        df = pd.read_excel(fp, sheet_name=sheet_name, engine='pyxlsb', header=None)
    except ValueError:
        return {}
    header_row = None
    for i in range(min(6, len(df))):
        if list(df.iloc[i, 0:4]) == ['Year', 'Type', 'Terms', 'Account']:
            header_row = i
            break
    if header_row is None:
        return {}
    week_cols = {}
    for col in range(4, df.shape[1]):
        v = df.iloc[header_row, col]
        try:
            wk = int(v)
        except (TypeError, ValueError):
            continue
        if 1 <= wk <= 52:
            week_cols[wk] = col
    total_row = None
    for i in range(header_row + 1, len(df)):
        acct = df.iloc[i, 3]
        if isinstance(acct, str) and acct.strip().upper().startswith('TOTAL'):
            total_row = i
            break
    if total_row is None:
        return {}
    yr = df.iloc[header_row + 1, 0]
    try:
        yr = int(yr)
    except (TypeError, ValueError):
        return {}
    weeks = [None] * 52
    for wk, col in week_cols.items():
        v = df.iloc[total_row, col]
        weeks[wk - 1] = float(v) if pd.notna(v) else None
    return {yr: weeks}

def latest_nonzero(weeks):
    """Scans a [wk1..52] array from the end and returns the last non-None,
    non-zero value -- "the current balance" convention for the hand-entered
    AP - Key / AP - CNS snapshots (see read_ap_vendor_total). Returns None
    if every week is None/zero."""
    for v in reversed(weeks or []):
        if v not in (None, 0, 0.0):
            return v
    return None

def js_arr_n(lst):
    """Like js_arr, but None -> JS null instead of the string 'None'."""
    return '[' + ','.join('null' if v is None else str(v) for v in lst) + ']'

def read_all():
    d = {}
    df_s  = pd.read_excel(FP, sheet_name='Summary',             engine='pyxlsb', header=None)
    df_a  = pd.read_excel(FP, sheet_name='2026 Actual',         engine='pyxlsb', header=None)
    df_b  = pd.read_excel(FP, sheet_name='2026 Budget',         engine='pyxlsb', header=None)
    df_25 = pd.read_excel(FP, sheet_name='2025 Actual',         engine='pyxlsb', header=None)
    df_cf = pd.read_excel(FP, sheet_name='Cash Flow - Tracker', engine='pyxlsb', header=None)
    df_inv= pd.read_excel(FP, sheet_name='Inventory',           engine='pyxlsb', header=None)
    df_oo = pd.read_excel(FP, sheet_name='On Order',            engine='pyxlsb', header=None)
    df_lb = pd.read_excel(FP, sheet_name='Labor',                engine='pyxlsb', header=None)
    df_pr = pd.read_excel(FP, sheet_name='Payroll',              engine='pyxlsb', header=None)

    d['cfw_history'] = read_cash_flow_weekly_history(FP)
    for yr, weeks in read_bank_balance_history(FP).items():
        d['cfw_history'].setdefault(yr, {})['bank_balance'] = weeks
    d['ar_history'] = read_ar_history(FP)
    d['ap_key_xl'] = read_ap_vendor_total(FP, 'AP - Key')
    d['ap_cns_xl'] = read_ap_vendor_total(FP, 'AP - CNS')

    def sc(row, col):
        r = row - 1
        c = {'B':1,'C':2,'D':3,'G':6,'H':7,'I':8}[col]
        v = df_s.iloc[r, c]
        return float(v) if pd.notna(v) else 0.0

    WK = int(df_s.iloc[2, 1])
    d['week']       = WK
    d['net_inc_var']= sc(21, 'D')
    d['ytd_status'] = 'ahead' if d['net_inc_var'] >= 0 else 'behind'
    for key, row in [('rev',11),('cogs',13),('opex',15),('net_income',21),
                      ('cash_in',68),('cash_out',75),('cf_var_pre',77),('cf_var_post',81),
                      ('credit_line',79)]:
        d[key] = {k: sc(row, c) for k,c in [('plan','B'),('act','C'),('var','D'),
                                              ('ann_plan','G'),('ann_proj','H'),('ann_var','I')]}

    def get_a(name):
        for i,row in df_a.iterrows():
            if str(row[0]).strip() == name: return float(row[WK]) if pd.notna(row[WK]) else 0.0
        return 0.0
    def get_b(name, wks=None):
        w = wks or WK
        for i,row in df_b.iterrows():
            if str(row[0]).strip() == name:
                return sum(float(row[j]) for j in range(1, w+1) if pd.notna(row[j]))
        return 0.0
    def get_r(name):
        for i,row in df_b.iterrows():
            if str(row[0]).strip() == name:
                return sum(float(row[j]) for j in range(WK+1, 53) if pd.notna(row[j]))
        return 0.0
    def get_ann(name):
        for i,row in df_b.iterrows():
            if str(row[0]).strip() == name: return float(row[59]) if pd.notna(row[59]) else 0.0
        return 0.0
    def proj(name): return get_a(name) + get_r(name)

    # Cash Bridge's COGS - A/P CNS / A/P - Key Vendors Plan: the 2026
    # Budget's remaining-weeks plan for the four line items those two
    # live-A/P feeds actually cover -- Spring/Summer, Fall/Winter, Special,
    # and Screen Printing Wholesale -- not Product Testing or any other
    # 2026 Budget COGS line, and not the broader Summary-sheet COGS total
    # used elsewhere on the dashboard.
    COGS_AP_PLAN_LINES = ['COGS - Spring/Summer', 'COGS - Fall/Winter', 'COGS - Special',
                           'COGS - Screen Printing Wholesale']
    d['cogs_ap_remaining_plan'] = sum(get_r(name) for name in COGS_AP_PLAN_LINES)

    rev_map = [
        ('Jettylife.com (DTC)',   'DTC - Jettylife.com'),
        ('Wholesale',             'Wholesale Revenue'),
        ('Screen printing (INK)', 'Screen Printing Revenue'),
        ('Flagship Store',        'DTC - Flagship Store'),
        ('Long Branch',           'DTC - Long Branch'),
        ('Mobile Store & Tent',   'DTC - Mobile Store & Tent'),
        ('JRF Screen Printing',   'JRF - Screen Printing'),
    ]
    d['rev_lines'] = []
    for lbl, k in rev_map:
        act, plan, ann_plan_v, ann_proj_v = get_a(k), get_b(k), get_ann(k), proj(k)
        trend_v = seasonal_ann_proj(k, act, ann_plan_v, WK)
        d['rev_lines'].append((lbl, act, plan, ann_plan_v, ann_proj_v, trend_v))
    # Total Revenue's trend projection is just each channel's own trend summed
    # — every channel has its own seasonality curve, so there's no single
    # curve for "all revenue" to run the ytd/expected-pace math against directly.
    d['rev_trend_total'] = sum(tp for *_, tp in d['rev_lines'])
    # Net Income's trend blends Revenue's seasonality-adjusted trend (a real,
    # validated signal -- consumer buying seasonality repeats year to year)
    # with COGS+Labor+Shipping's, OpEx's, and EBIT Expenses' own plan-paced
    # Full Year Projection, *not* a cost-side seasonality curve -- we tried
    # that and reverted it, since bill/payment entry timing doesn't repeat
    # the way revenue does (see the note by COGS_LS_TOTAL above). Building
    # off net_income['ann_proj'] rather than re-deriving from rev/cogs/opex
    # keeps EBIT Expenses (interest, D&A, taxes -- read directly off the
    # Summary sheet, not broken out into `d`) correctly in the total: only
    # the revenue term is swapped for its trend-adjusted counterpart.
    d['net_income_trend'] = d['net_income']['ann_proj'] + (d['rev_trend_total'] - d['rev']['ann_proj'])

    # ── 52-week trend charts (Summary tab) ──────────────────────────────────
    # Row indices are consistent across '2025 Actual', '2026 Budget', and
    # '2026 Actual' -- all three sheets share the same P&L layout.
    TREND_ROWS = {'revenue': 18, 'cogs': 32, 'opex': 96, 'ebit': 105, 'net_income': 107}
    # 2025's Total COGS/OpEx/EBIT rows are SUM formulas over mostly-blank
    # category rows -- only the 12 month-end weeks (where we backfilled a
    # typed-over cumulative value, per the Net Income trend project) hold a
    # real figure; every other week reads as a plain 0, not blank, so it has
    # to be filtered explicitly rather than relying on NaN. Net Income is
    # itself a formula off those two (Income - COGS - OpEx), so at every
    # non-month-end week it silently evaluates to Income - 0 - 0 == Income --
    # caught by cross-checking it against the Total Income row directly.
    # Revenue is the one row genuinely populated for all 52 weeks.
    MONTH_END_WEEKS = {4,8,13,17,21,26,30,34,39,43,47,52}

    def cum_from_weekly_budget(df, row):
        weekly = [df.iloc[row, c] or 0 for c in range(1, 53)]
        out = []; s = 0
        for v in weekly:
            s += v
            out.append(s)
        return out  # 52 values, index 0 = week 1

    def cum_from_actual(df, row, thru_wk=52, month_end_only=False):
        out = []
        for wk in range(1, 53):
            v = df.iloc[row, wk]
            if wk > thru_wk or pd.isna(v) or (month_end_only and wk not in MONTH_END_WEEKS):
                out.append(None)
            else:
                out.append(float(v))
        return out

    d['trend_charts'] = {}
    for key, row in TREND_ROWS.items():
        d['trend_charts'][key] = {
            'act_2025':  cum_from_actual(df_25, row, month_end_only=(key in ('cogs','opex','ebit','net_income'))),
            'plan_2026': cum_from_weekly_budget(df_b, row),
            'act_2026':  cum_from_actual(df_a, row, thru_wk=WK),
        }

    # Trend continuation for weeks WK..52, anchored to the real actual value
    # at WK so it joins the solid actual line without a visible seam:
    #  - revenue: each channel's own validated seasonality curve, summed --
    #    this naturally lands on the real actual total at week WK by
    #    construction (trend_v * cum_weight_at_WK == that channel's actual).
    #  - cogs/opex/ebit: parallels the remaining *budgeted* weekly shape --
    #    no cost-side seasonality curve exists (see the note above
    #    CHANNEL_SEASONALITY for why that was tried and reverted).
    #  - net_income: revenue trend minus cogs/opex/ebit trend, week by week.
    # Normalized so each channel's own cumulative weight hits exactly 1.0 at
    # week 52 (raw monthly weights are rounded to 3 decimals and sum to
    # ~0.998-1.001 per channel) -- without this, the week-52 chart point
    # drifts a fraction of a percent off the "Trend Proj." figure already
    # shown on the Total Revenue card, which uses trend_v as-is.
    def _norm_cum_weights(k):
        raw = _weekly_weights(CHANNEL_SEASONALITY[k])
        total = sum(raw)
        return [sum(raw[:i]) / total for i in range(1, 53)]
    rev_cum_weights = {k: _norm_cum_weights(k) for _, k in rev_map}
    rev_trend_v = {k: seasonal_ann_proj(k, get_a(k), get_ann(k), WK) for _, k in rev_map}
    rev_trend = [None]*52
    for wk in range(WK, 53):
        rev_trend[wk-1] = sum(rev_trend_v[k] * rev_cum_weights[k][wk-1] for _, k in rev_map)
    d['trend_charts']['revenue']['trend_2026'] = rev_trend

    def plan_paced_trend(tc):
        plan = tc['plan_2026']; act_wk = tc['act_2026'][WK-1]
        return [None]*(WK-1) + [act_wk] + [act_wk + (plan[wk-1]-plan[WK-1]) for wk in range(WK+1, 53)]

    for key in ('cogs', 'opex', 'ebit'):
        d['trend_charts'][key]['trend_2026'] = plan_paced_trend(d['trend_charts'][key])

    d['trend_charts']['net_income']['trend_2026'] = [
        None if any(x is None for x in (r,c,o,e)) else r-c-o-e
        for r,c,o,e in zip(d['trend_charts']['revenue']['trend_2026'],
                            d['trend_charts']['cogs']['trend_2026'],
                            d['trend_charts']['opex']['trend_2026'],
                            d['trend_charts']['ebit']['trend_2026'])
    ]

    cogs_map = [
        ('Contract design — brand', 'Contract Design - Brand'),
        ('Contract design — INK',   'Contract Design - Ink'),
        ('Shipping — brand',        'Shipping - BRAND'),
        ('Shipping — INK',          'Shipping - INK'),
    ]
    d['cogs_lines'] = [(lbl, get_a(k), get_b(k), get_ann(k), proj(k)) for lbl,k in cogs_map]

    opex_map = [
        ('Advertising',                  'Advertising'),
        ('Advertising — INK',            'Advertising - Ink'),
        ('Ambassador & influencer',      'Ambassador & Influencer Expenses'),
        ('Box truck',                    'Box Truck'),
        ('Dues & subscriptions',         'Dues & Subscriptions'),
        ('Equipment & repairs',          'Equipment & Repairs'),
        ('Ford Transit — INK van',       'Ford Transit - Ink Van Expenses'),
        ('INK merch store profit share', 'Ink Merch Store Profit Share'),
        ('MKG — bags',                   'MKG - Bags'),
        ('MKG — banners, POP & misc.',   'MKG - Banners, P.O.P & Misc.'),
        ('MKG — catalogs',               'MKG - Catalogs'),
        ('MKG — direct mail',            'MKG - Direct Mail'),
        ('MKG — hang tags & size tags',  'MKG - Hang Tags & Size Tags'),
        ('MKG — photography',            'MKG - Photography'),
        ('MKG — stickers',               'MKG - Stickers'),
        ('MKG — tent sale expenses',     'MKG - Tent Sale Expenses'),
        ('MKG — trade show',             'MKG - Trade Show Expense'),
        ('MKG — video',                  'MKG - Video'),
        ('Miscellaneous warehouse exp.', 'Miscellaneous Warehouse Expenses'),
        ('Office supplies',              'Office Expense - Office Supplies'),
        ('Payroll service',              'Payroll Service'),
        ('Permits — 176 E Bay',          'Permits & Building - 176 E Bay'),
        ('Permits — 700 S Main',         'Permits & Building - 700 S Main St'),
        ('Permits — Flagship',           'Permits & Building - Flagship Store'),
        ('Permits — Jetty INK',          'Permits & Building - Jetty Ink'),
        ('Permits — Long Branch',        'Permits & Building - Long Branch'),
        ('Screen printing supplies',     'Screen Printing Supplies'),
        ('Selling exp. — rep supplies',  'Selling Expenses - Rep Supplies'),
        ('Selling exp. — rep travel',    'Selling Expenses - Rep Travel'),
        ('Shipping supplies',            'Shipping Supplies'),
        ('Software & inventory control', 'Software & Inventory Control'),
        ('Team building',                'Team Building'),
        ('Website expenses',             'Website Expenses'),
    ]
    opex_categories = read_categories()
    if opex_categories:
        # Categories tab is the authoritative full line-item list once it exists —
        # it covers every real OpEx row, not just the hand-picked opex_map subset.
        label_overrides = {k: lbl for lbl, k in opex_map}
        opex_keys = list(opex_categories.keys())
        opex_lines = [(label_overrides.get(k, k), get_a(k), get_b(k), get_ann(k), proj(k)) for k in opex_keys]
    else:
        opex_keys = [k for lbl, k in opex_map]
        opex_lines = [(lbl, get_a(k), get_b(k), get_ann(k), proj(k)) for lbl, k in opex_map]
    d['opex_lines'] = opex_lines
    d['opex_keys'] = opex_keys
    d['opex_categories'] = opex_categories

    # Cash flow
    cin_w,cin_d,cin_i,cin_t,cout_b,cout_i,cout_o,cout_l,cout_e = [],[],[],[],[],[],[],[],[]
    proj_w,proj_d,proj_i,proj_b,proj_i2,proj_o,proj_l,proj_e  = [],[],[],[],[],[],[],[]
    rw=rd=ri=rb=ri2=ro=rl=re=0
    labels=[]
    # Weekly cash-in schedule: the sheet's "Actual" columns (10-13) are cumulative,
    # so per-week cash in is the delta vs the prior week's cumulative total. Plan
    # columns (5-8) are already per-week, so weeks beyond WK can be shown as-is
    # for a forecast of the remaining schedule.
    cf_weekly, cf_forecast = [], []
    prev_w = prev_d = prev_i = prev_t = 0.0
    # Full-year plan totals (all 52 weeks), for an annual projection that's
    # consistent with the actual+remaining-plan convention used everywhere
    # else: ann_proj = ytd_act + remaining_plan, so ann_var always equals
    # ytd_var exactly (remaining_plan cancels out of the subtraction).
    aw_full=ad_full=ai_full=ab_full=ai2_full=ao_full=al_full=ae_full=0.0
    for _,row in df_cf.iterrows():
        if pd.notna(row[2]) and str(row[2]) != 'Week':
            try: wk = int(row[2])
            except: continue
            if wk > 52: break
            plan_w = float(row[5]) if pd.notna(row[5]) else 0.0
            plan_d = float(row[6]) if pd.notna(row[6]) else 0.0
            plan_i = float(row[7]) if pd.notna(row[7]) else 0.0
            plan_t = float(row[8]) if pd.notna(row[8]) else (plan_w+plan_d+plan_i)
            cout_b_plan = float(row[15]) if pd.notna(row[15]) else 0.0
            cout_i_plan = float(row[16]) if pd.notna(row[16]) else 0.0
            cout_o_plan = float(row[17]) if pd.notna(row[17]) else 0.0
            cout_l_plan = float(row[18]) if pd.notna(row[18]) else 0.0
            cout_e_plan = float(row[19]) if pd.notna(row[19]) else 0.0
            aw_full+=plan_w; ad_full+=plan_d; ai_full+=plan_i
            ab_full+=cout_b_plan; ai2_full+=cout_i_plan; ao_full+=cout_o_plan; al_full+=cout_l_plan; ae_full+=cout_e_plan
            if wk > WK:
                cf_forecast.append(dict(wk=wk, w=plan_w, d=plan_d, i=plan_i, t=plan_t))
                continue
            rw  += plan_w
            rd  += plan_d
            ri  += plan_i
            rb  += cout_b_plan
            ri2 += cout_i_plan
            ro  += cout_o_plan
            rl  += cout_l_plan
            re  += cout_e_plan
            act  = float(row[13]) if pd.notna(row[13]) else 0
            if act > 0:
                cum_w = float(row[10]) if pd.notna(row[10]) else prev_w
                cum_d = float(row[11]) if pd.notna(row[11]) else prev_d
                cum_i = float(row[12]) if pd.notna(row[12]) else prev_i
                cum_t = act
                cf_weekly.append(dict(wk=wk, w=cum_w-prev_w, d=cum_d-prev_d, i=cum_i-prev_i, t=cum_t-prev_t,
                                       pw=plan_w, pd_=plan_d, pi=plan_i, pt=plan_t))
                prev_w, prev_d, prev_i, prev_t = cum_w, cum_d, cum_i, cum_t

                labels.append("Wk " + str(wk))
                cin_w.append(round(cum_w)); cin_d.append(round(cum_d)); cin_i.append(round(cum_i))
                cin_t.append(round(cum_t))
                cout_b.append(round(float(row[21]) if pd.notna(row[21]) else 0))
                cout_i.append(round(float(row[22]) if pd.notna(row[22]) else 0))
                cout_o.append(round(float(row[23]) if pd.notna(row[23]) else 0))
                cout_l.append(round(float(row[24]) if pd.notna(row[24]) else 0))
                cout_e.append(round(float(row[25]) if pd.notna(row[25]) else 0))
                proj_w.append(round(rw));  proj_d.append(round(rd));  proj_i.append(round(ri))
                proj_b.append(round(rb));  proj_i2.append(round(ri2)); proj_o.append(round(ro))
                proj_l.append(round(rl));  proj_e.append(round(re))

    d['cf'] = dict(labels=labels,
        cin_w=cin_w, cin_d=cin_d, cin_i=cin_i, cin_t=cin_t,
        cout_b=cout_b, cout_i=cout_i, cout_o=cout_o, cout_l=cout_l, cout_e=cout_e,
        proj_w=proj_w, proj_d=proj_d, proj_i=proj_i,
        proj_b=proj_b, proj_i2=proj_i2, proj_o=proj_o, proj_l=proj_l, proj_e=proj_e)
    d['cf_ann_plan'] = dict(w=aw_full, d=ad_full, i=ai_full,
                             b=ab_full, i2=ai2_full, o=ao_full, l=al_full, e=ae_full)
    d['cf_weekly']   = cf_weekly
    d['cf_forecast'] = cf_forecast

    # ── Bank Position & Weekly Reconciliation (Cash Flow - Weekly / Bank Balance) ──
    # Added once actual bank activity started getting tracked directly: "Cash
    # Flow - Weekly" is a straight pull of Xero's General Ledger Report --
    # every real credit/debit across both bank accounts (Columbia + BOA) for
    # the week, no plan-vs-channel modeling involved. "Bank Balance" is the
    # two accounts' actual recorded starting/ending balances, independently
    # entered from online banking. Because these come from two different
    # sources, "prior week's actual balance + this week's true cash in - true
    # cash out" should equal this week's recorded actual balance -- any gap
    # is a real unreconciled transaction or double-entry, not a modeling
    # artifact, which is exactly why this is useful as an ongoing weekly
    # check rather than just a one-time cleanup.
    # These two tabs have come and gone before (rebuilt from scratch, or lost
    # entirely to an unsaved-sheet mishap) -- a missing sheet should degrade
    # this section to empty, not take down the whole build.
    try:
        df_bb  = pd.read_excel(FP, sheet_name='Bank Balance',      engine='pyxlsb', header=None)
    except ValueError as e:
        print("Bank Position skipped, sheet not found:", e)
        df_bb = None

    # Cash Flow - Weekly's own Year-blocked layout (Year/Account/wk1..wk52)
    # is already parsed correctly by read_cash_flow_weekly_history() into
    # d['cfw_history'] -- the old cfw_row() here predated that layout (it
    # matched row[0] against the label directly, with weeks starting at
    # column 1), so it silently matched nothing once the sheet gained a
    # separate Year column, and this whole section went quietly empty.
    # Pulling from cfw_history for the latest year sidesteps that bug
    # entirely instead of patching the old row-scan.
    cfw_hist = d.get('cfw_history') or {}
    cfw_year = max(cfw_hist.keys()) if cfw_hist else None
    cfw_cur = cfw_hist.get(cfw_year) if cfw_year else {}
    r_cin, r_cout, r_draw, r_pay = (cfw_cur.get(k) for k in ('cash_in', 'cash_out', 'cl_draw', 'cl_paydown'))

    if df_bb is not None:
        bb_2026 = df_bb[df_bb[0] == 2026]
        bb_total_by_wk = {int(row[2]): float(row[8]) for _, row in bb_2026.iterrows() if pd.notna(row[8]) and row[8] != 0}
        bb_accts_by_wk = {int(row[2]): dict(columbia=float(row[5]) if pd.notna(row[5]) else 0.0,
                                             boa=float(row[6]) if pd.notna(row[6]) else 0.0,
                                             ramp=float(row[7]) if pd.notna(row[7]) else 0.0)
                          for _, row in bb_2026.iterrows() if pd.notna(row[8]) and row[8] != 0}
        prior_year_row = df_bb[(df_bb[0] == 2025) & (df_bb[2] == 52)]
        bb_prior_year_end = float(prior_year_row.iloc[0][8]) if len(prior_year_row) else None
    else:
        bb_total_by_wk, bb_accts_by_wk, bb_prior_year_end = {}, {}, None

    bank_position = []
    prev_end = bb_prior_year_end
    if r_cin is not None and r_cout is not None and prev_end is not None:
        for wk in range(1, 53):
            cin  = r_cin[wk - 1]
            cout = r_cout[wk - 1]
            if cin is None or cout is None:
                break  # no further weeks entered yet
            draw = r_draw[wk - 1] if r_draw is not None and r_draw[wk - 1] is not None else 0.0
            pay  = r_pay[wk - 1]  if r_pay  is not None and r_pay[wk - 1]  is not None else 0.0
            computed_end = prev_end + cin - cout
            actual_end = bb_total_by_wk.get(wk)
            gap = (actual_end - computed_end) if actual_end is not None else None
            bank_position.append(dict(wk=wk, start=prev_end, cash_in=cin, cash_out=cout,
                                       cl_draw=draw, cl_paydown=pay, computed_end=computed_end,
                                       actual_end=actual_end, gap=gap, accts=bb_accts_by_wk.get(wk)))
            # Chain off the actual recorded balance when we have one, so a
            # single week's gap doesn't compound into every week after it.
            prev_end = actual_end if actual_end is not None else computed_end

    d['bank_position']  = bank_position
    d['bp_latest']       = bank_position[-1] if bank_position else None
    d['bp_year_start']   = bb_prior_year_end
    GAP_WATCH = 15000  # weekly reconciliation gap large enough to flag for follow-up
    d['bp_gap_watch']    = GAP_WATCH
    d['bp_flags']        = [bp for bp in bank_position if bp['gap'] is not None and abs(bp['gap']) > GAP_WATCH]
    if bank_position:
        d['bp_ytd'] = dict(
            cash_in    = sum(bp['cash_in']    for bp in bank_position),
            cash_out   = sum(bp['cash_out']   for bp in bank_position),
            cl_draw    = sum(bp['cl_draw']    for bp in bank_position),
            cl_paydown = sum(bp['cl_paydown'] for bp in bank_position),
        )
        d['bp_ytd']['net'] = d['bp_ytd']['cash_in'] - d['bp_ytd']['cash_out']
    else:
        d['bp_ytd'] = None

    # ── Accounts Receivable (AR tab) ────────────────────────────────────────
    # Weekly aging, hand-entered from Xero's Accounts Receivable Aging report
    # for both entities. Total (0-90 Days) per entity is real, already-
    # invoiced money expected to convert to cash soon -- concrete in a way a
    # pace extrapolation isn't, which is why the Cash Bridge Trend below
    # pulls it in. A missing/rebuilt sheet degrades to no A/R signal rather
    # than breaking the build, same convention as Bank Position above.
    try:
        df_ar = pd.read_excel(FP, sheet_name='AR', engine='pyxlsb', header=None)
    except ValueError as e:
        print("AR aging skipped, sheet not found:", e)
        df_ar = None

    ar_weekly = []
    if df_ar is not None:
        for _, row in df_ar[df_ar[0] == 2026].iterrows():
            if not pd.notna(row[2]):
                continue
            brand_090 = float(row[10]) if pd.notna(row[10]) else 0.0
            ink_090   = float(row[17]) if pd.notna(row[17]) else 0.0
            if not brand_090 and not ink_090:
                continue  # week not entered yet
            ar_weekly.append(dict(wk=int(row[2]), brand_090=brand_090, ink_090=ink_090))
    d['ar_weekly'] = ar_weekly
    d['ar_latest'] = ar_weekly[-1] if ar_weekly else None

    # Credit line: monthly cadence (unlike the weekly cash schedule above),
    # from two different sources -- the bank's monthly Borrowing Certificate
    # (actual balance, submitted after the fact) for the 2025 and 2026
    # actual trajectories, and the original "Cash Flow - Plan" sheet for the
    # full-year 2026 plan. That plan sheet is never edited after the year
    # starts, so it stays a true fixed reference rather than drifting to
    # match reality the way the certificate's actuals naturally do.
    MONTH_ORDER = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
    df_cert = pd.read_excel(FP, sheet_name='Columbia Certificate', engine='pyxlsb', header=None)
    cert_2025, cert_2026, loc_status = {}, {}, None
    for r in range(6, df_cert.shape[0]):
        row = df_cert.iloc[r]
        yr, mon, bal = row[0], row[1], row[7]
        if pd.isna(yr) or pd.isna(mon) or pd.isna(bal):
            continue  # future months not yet certified -- placeholder row
        entry = dict(balance=float(bal), month=str(mon).strip(),
                     coverage_x=(float(row[11]) if pd.notna(row[11]) and not isinstance(row[11], str) else None),
                     status=(str(row[13]) if pd.notna(row[13]) else ''))
        (cert_2025 if int(yr) == 2025 else cert_2026)[str(mon).strip()] = entry
        if int(yr) == 2026:
            loc_status = entry  # rows are chronological -- last write is the latest certified month
    d['loc_2025']     = [cert_2025[m]['balance'] if m in cert_2025 else None for m in MONTH_ORDER]
    d['loc_2026_act'] = [cert_2026[m]['balance'] if m in cert_2026 else None for m in MONTH_ORDER]
    d['loc_status']   = loc_status

    df_locplan = pd.read_excel(FP, sheet_name='Cash Flow - Plan', engine='pyxlsb', header=None)
    plan_month_cols = [1,2,3,4,5,6,8,9,10,11,12,13]  # Jan..Jun, Jul..Dec (skips the H1/H2 rollup columns)
    def loc_plan_row(r):
        row = df_locplan.iloc[r]
        return [float(row[c]) if pd.notna(row[c]) else 0.0 for c in plan_month_cols]
    loc_plan_bom = loc_plan_row(34)  # "Balance BOM"
    loc_plan_use = loc_plan_row(35)  # "Use of Credit"
    loc_plan_pay = loc_plan_row(36)  # "Paydown of Credit"
    loc_plan_eom = loc_plan_row(37)  # "Balance EOM"
    d['loc_plan_bom']     = loc_plan_bom
    d['loc_plan_monthly'] = [dict(month=MONTH_ORDER[i], bom=loc_plan_bom[i], draw=loc_plan_use[i],
                                   paydown=loc_plan_pay[i], eom=loc_plan_eom[i]) for i in range(12)]

    # Trend continuation: certified actual through the latest certified
    # month, then the original plan's remaining months -- anchored on the
    # last actual value (duplicated at the seam) so the dashed line picks up
    # exactly where the solid actual line stops, same convention as the
    # Summary tab's 52-week trend charts.
    n_elapsed = sum(1 for v in d['loc_2026_act'] if v is not None)
    d['loc_n_elapsed'] = n_elapsed
    if n_elapsed == 0:
        d['loc_2026_trend'] = list(loc_plan_bom)
    else:
        d['loc_2026_trend'] = ([None]*(n_elapsed-1) + [d['loc_2026_act'][n_elapsed-1]]
                                + loc_plan_bom[n_elapsed:12])

    d['ap_cns'] = fetch_ap_cns()
    d['ap_key_vendors_weekly'] = fetch_ap_key_vendors_weekly()
    d['ap_key_vendors_terms']  = fetch_ap_key_vendors_terms()
    d['ap_key_vendors_open']   = fetch_ap_key_vendors_open()

    # Inventory. A missing weekly reading is not the same as zero on hand — it
    # just means nobody recorded a count that week. carry_forward fills those
    # gaps with the last known value (instead of a fake drop to $0/0 units)
    # and flags which weeks were filled in, so charts can mark them distinctly.
    def carry_forward(pairs):
        values = {}; estimated = set(); prev = None
        for wk, v in sorted(pairs):
            if v is not None:
                values[wk] = v; prev = v
            elif prev is not None:
                values[wk] = prev; estimated.add(wk)
        return values, estimated

    d['inv_est'] = {}  # series name -> set of weeks whose value was carried forward

    # The sheet's Inventory tab was restructured to a simpler layout: one
    # Total INV COG (col 5) / Total Units (col 6) pair per week, plus six
    # named sell-through columns (S24/F24/S25/F25/S26/F26, cols 8-13) — no
    # more per-category COG/units breakdown. Col 7 (Avg Cost/Unit) is a
    # formula-error placeholder in the export and is ignored; we compute the
    # average ourselves from COG/units instead.
    inv_wks = []
    pairs_c = []; pairs_u = []
    for _,row in df_inv[df_inv[0]==2026].iterrows():
        if not pd.notna(row[2]): continue
        wk = int(row[2])
        if wk > WK: continue
        inv_wks.append(wk)
        c = float(row[5]) if pd.notna(row[5]) and isinstance(row[5],(int,float)) else None
        u = float(row[6]) if pd.notna(row[6]) and isinstance(row[6],(int,float)) else None
        pairs_c.append((wk,c)); pairs_u.append((wk,u))

    cog_v, cog_e = carry_forward(pairs_c)
    u_v,   u_e   = carry_forward(pairs_u)
    d['inv_est']['cog_2026']   = cog_e
    d['inv_est']['units_2026'] = u_e
    d['cog_2026']   = {wk: round(cog_v[wk]) for wk in inv_wks if wk in cog_v}
    d['units_2026'] = {wk: round(u_v[wk])   for wk in inv_wks if wk in u_v}
    d['inv_wks']      = inv_wks
    d['inv_latest_wk']= inv_wks[-1] if inv_wks else None
    d['inv_prev_wk']  = inv_wks[-2] if len(inv_wks) > 1 else None

    for yr,ck,uk in [(2024,'cog_2024','units_2024'),(2025,'cog_2025','units_2025')]:
        pairs_c=[]; pairs_u=[]
        for _,row in df_inv[df_inv[0]==yr].iterrows():
            if pd.notna(row[2]):
                try:
                    wk=int(row[2])
                    cv = float(row[5]) if pd.notna(row[5]) and isinstance(row[5],(int,float)) else None
                    uv = float(row[6]) if pd.notna(row[6]) and isinstance(row[6],(int,float)) else None
                    pairs_c.append((wk,cv)); pairs_u.append((wk,uv))
                except: pass
        cog_v, cog_e = carry_forward(pairs_c)
        u_v,   u_e   = carry_forward(pairs_u)
        d[ck] = {wk: round(v) for wk,v in cog_v.items()}
        d[uk] = {wk: round(v) for wk,v in u_v.items()}
        d['inv_est'][ck] = cog_e
        d['inv_est'][uk] = u_e

    lw = d['inv_latest_wk']
    d['yoy_cog_2025']=0; d['yoy_units_2025']=0
    if lw:
        for _,row in df_inv[df_inv[0]==2025].iterrows():
            if pd.notna(row[2]) and int(row[2])==lw:
                d['yoy_cog_2025']  = round(float(row[5])) if pd.notna(row[5]) else 0
                d['yoy_units_2025']= round(float(row[6])) if pd.notna(row[6]) else 0
                break

    # Sell-through: each season has its own dedicated column now (8-13), but
    # every column is still only populated in the one year that season was
    # actively selling (e.g. F25 sell-through is tracked in 2026 rows, since
    # that's when F25 units were on the floor) — pull each from its actual
    # active year, not from the year its name suggests.
    def extract_st(year, col):
        pairs = []
        for _,row in df_inv[df_inv[0]==year].iterrows():
            if pd.notna(row[2]):
                wk = int(row[2])
                v = float(row[col]) if pd.notna(row[col]) and isinstance(row[col],(int,float)) and 0<float(row[col])<=1 else None
                pairs.append((wk, round(v*100,1) if v is not None else None))
        return carry_forward(pairs)

    d['st_s24'], d['inv_est']['st_s24'] = extract_st(2024, 8)
    d['st_f24'], d['inv_est']['st_f24'] = extract_st(2024, 9)
    d['st_s25'], d['inv_est']['st_s25'] = extract_st(2025, 10)
    d['st_f25'], d['inv_est']['st_f25'] = extract_st(2026, 11)
    d['st_s26'], d['inv_est']['st_s26'] = extract_st(2026, 12)
    d['st_f26'], d['inv_est']['st_f26'] = extract_st(2026, 13)

    # On Order: total, by product category, and by half — for both this year
    # and next, since orders for 2027 H1 start landing well before 2026 closes.
    # Column layout (row 4 headers): 5=Total, 6-13=categories (Clearance,
    # Collab, F26, S25, S26, SP27, SU27, STARBOARD), 14=H1 2026, 15=H2 2026,
    # 16=H1 2027, 17=H2 2027. STARBOARD was added as a category after this
    # was first written, shifting the H1/H2 columns one to the right of what
    # the code assumed -- verified against the sheet's own Total (col 5,
    # which equals the sum of cols 6-13, and separately equals the sum of
    # cols 14-17) that 14/15/16/17 is the correct mapping, not 13/14/15/16.
    d['on_order']={'total':0,'h1':0,'h2':0}
    d['on_order_cats'] = []
    d['on_order_periods'] = {'h1_2026':0,'h2_2026':0,'h1_2027':0,'h2_2027':0}
    for _,row in df_oo.iterrows():
        if pd.notna(row[0]) and row[0]==2026 and pd.notna(row[2]) and int(row[2])==WK:
            d['on_order']={
                'total': float(row[5])  if pd.notna(row[5])  else 0,
                'h1':    float(row[14]) if pd.notna(row[14]) else 0,
                'h2':    float(row[15]) if pd.notna(row[15]) else 0,
            }
            cat_cols_oo = [('Clearance',6),('Collab',7),('F26',8),('S25',9),('S26',10),('SP27',11),('SU27',12),('STARBOARD',13)]
            d['on_order_cats'] = [(name, float(row[c]) if pd.notna(row[c]) else 0.0) for name,c in cat_cols_oo]
            d['on_order_periods'] = {
                'h1_2026': float(row[14]) if pd.notna(row[14]) else 0.0,
                'h2_2026': float(row[15]) if pd.notna(row[15]) else 0.0,
                'h1_2027': float(row[16]) if pd.notna(row[16]) else 0.0,
                'h2_2027': float(row[17]) if pd.notna(row[17]) else 0.0,
            }
            break

    def get_a_at(name, wk):
        for _,row in df_a.iterrows():
            if str(row[0]).strip() == name:
                return float(row[wk]) if pd.notna(row[wk]) else 0.0
        return 0.0

    for _,row in df_b.iterrows():
        if str(row[0])=='Wholesale Revenue':
            d['whsl_h1_plan'] = sum(float(row[j]) for j in range(1,27)  if pd.notna(row[j]))
            d['whsl_h2_plan'] = sum(float(row[j]) for j in range(27,53) if pd.notna(row[j]))
            break

    # Actual sheet columns are cumulative-to-date, so week 26 (the last week
    # of H1) gives a real H1-only actual regardless of which half we're in
    # now — no more "weeks left in H1" math that goes negative once H2 starts.
    whsl_h1_actual = get_a_at('Wholesale Revenue', min(WK, 26))
    whsl_ytd_actual = get_a_at('Wholesale Revenue', WK)
    whsl_h2_actual = max(0.0, whsl_ytd_actual - whsl_h1_actual) if WK > 26 else 0.0
    d['whsl_act']      = whsl_ytd_actual
    d['whsl_h1_actual']= whsl_h1_actual
    d['whsl_h2_actual']= whsl_h2_actual
    d['whsl_h1_proj']  = whsl_h1_actual + d['on_order']['h1']
    d['whsl_h2_proj']  = whsl_h2_actual + d['on_order']['h2']
    d['whsl_h1_gap']   = d['whsl_h1_plan'] - d['whsl_h1_proj']
    d['whsl_h2_gap']   = d['whsl_h2_plan'] - d['whsl_h2_proj']

    # Labor: monthly Plan/Actual/Variance for JETTY INK (cols 0-6) and JETTY BRAND
    # (cols 8-14). A month with no Actual yet hasn't closed — its Plan counts toward
    # the remaining-plan total (and so the full-year projection) but not YTD actual,
    # matching the actual+remaining-plan projection convention used elsewhere.
    def labor_year(col0, year):
        months=[]; ytd_act=ytd_plan=rem_plan=ann_plan=0.0
        for _,row in df_lb.iterrows():
            y = row[col0]
            if not isinstance(y, (int, float)) or pd.isna(y) or int(y) != year: continue
            plan = float(row[col0+4]) if pd.notna(row[col0+4]) else 0.0
            act  = row[col0+5]
            mon  = str(row[col0+1]).strip() if pd.notna(row[col0+1]) else ''
            ann_plan += plan
            if pd.notna(act):
                ytd_act += float(act); ytd_plan += plan
                months.append((mon, plan, float(act)))
            else:
                rem_plan += plan
                months.append((mon, plan, None))
        return dict(ytd_act=ytd_act, ytd_plan=ytd_plan, ytd_var=ytd_act-ytd_plan,
                    ann_plan=ann_plan, ann_proj=ytd_act+rem_plan,
                    ann_var=(ytd_act+rem_plan)-ann_plan, months=months)

    labor_ink   = labor_year(0, 2026)
    labor_brand = labor_year(8, 2026)
    labor_total = dict(
        ytd_act = labor_ink['ytd_act']+labor_brand['ytd_act'],
        ytd_plan= labor_ink['ytd_plan']+labor_brand['ytd_plan'],
        ann_plan= labor_ink['ann_plan']+labor_brand['ann_plan'],
        ann_proj= labor_ink['ann_proj']+labor_brand['ann_proj'],
    )
    labor_total['ytd_var'] = labor_total['ytd_act']-labor_total['ytd_plan']
    labor_total['ann_var'] = labor_total['ann_proj']-labor_total['ann_plan']
    labor_total['months']  = [(m, pi+pb, (ai+ab) if ai is not None and ab is not None else None)
                               for (m,pi,ai),(_,pb,ab) in zip(labor_ink['months'], labor_brand['months'])]
    d['labor_ink']   = labor_ink
    d['labor_brand'] = labor_brand
    d['labor_total'] = labor_total

    # Payroll: biweekly headcount + overtime. Weeks with no Total Payroll are the
    # off week in each pay period and carry no snapshot.
    payroll = []
    for _,row in df_pr.iterrows():
        if isinstance(row[0], (int, float)) and pd.notna(row[0]) and isinstance(row[4], (int, float)) and pd.notna(row[4]):
            pay_date = (pd.to_datetime(row[3], unit='D', origin='1899-12-30').strftime('%b %d, %Y')
                        if pd.notna(row[3]) and isinstance(row[3], (int, float)) else '')
            payroll.append(dict(
                wk=int(row[0]), pay_date=pay_date, total_payroll=float(row[4]),
                taxes=float(row[5]) if pd.notna(row[5]) else 0.0,
                net=float(row[6]) if pd.notna(row[6]) else 0.0,
                total=int(row[7]), pt=int(row[8]), ft=int(row[9]),
                hourly=int(row[10]), salary=int(row[11]),
                female=int(row[12]), male=int(row[13]),
                ink=int(row[14]), brand=int(row[15]),
                depts=dict(
                    ink_design=int(row[16]), ink_ops=int(row[17]), ink_prod=int(row[18]), ink_sales=int(row[19]),
                    logistics=int(row[20]), mkg=int(row[21]), prodev=int(row[22]), flagship=int(row[23]),
                    long_branch=int(row[24]), box_truck=int(row[25]), whsl_sales=int(row[26]), finance=int(row[27]),
                ),
                ot_hours=float(row[28]) if pd.notna(row[28]) else 0.0,
                ot_amount=float(row[29]) if pd.notna(row[29]) else 0.0,
            ))
    d['payroll']        = payroll
    d['payroll_latest'] = payroll[-1] if payroll else None

    return d

# Two known source-sheet labeling inconsistencies for the OpEx "Program
# Services" line, found by cross-checking against Summary's own Program
# Services figures: Monthly/Annual - Budget mistakenly tag that line's
# category as "Donation"; Actual (Cumulative) - Sorted spells it "Program
# Service" (singular) instead of "Program Services". Both get normalized
# to the same key so the line groups correctly regardless of source sheet.
JRF_OPEX_CATEGORY_FIX = {'Donation': 'Program Services', 'Program Service': 'Program Services'}

def _jrf_as_of_month(raw):
    """The 'As of' cell has been typed as plain text ('As of 7/31/26') and,
    after the user re-entered it, as a real date -- handle both so a future
    formatting change doesn't silently break the month lookup."""
    if isinstance(raw, datetime.datetime):
        return raw.month, raw.strftime('%m/%d/%y').lstrip('0').replace('/0', '/')
    s = str(raw).replace('As of', '').strip()
    month = int(s.split('/')[0])
    return month, s

def _jrf_section_rows(ws, cat_col, acct_col, val_fn):
    """Yields (section, category, account, values) for every real line-item
    row in a sheet shaped like Monthly - Budget / Actual (Cumulative) -
    Sorted: an Income/Cost of Goods Sold/Operating Expenses section header,
    then Category+Account rows underneath, with 'Total X' subtotal rows to
    skip. Section-scoped because category labels like 'Event' and 'Other'
    repeat across the Revenue and OpEx sections."""
    section = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        cat, acct = row[cat_col], row[acct_col]
        if cat is None and acct in ('Income', 'Cost of Goods Sold', 'Operating Expenses'):
            section = {'Income': 'rev', 'Cost of Goods Sold': 'don', 'Operating Expenses': 'opex'}[acct]
            continue
        if cat is None or acct is None: continue
        if str(acct).strip().startswith('Total') or str(acct).strip() == 'Gross Profit': continue
        yield section, str(cat).strip(), str(acct).strip(), val_fn(row)

def read_jrf():
    """JRF (Jetty Rock Foundation) tracks its own budget in a separate,
    simpler workbook (monthly cadence, no weekly detail) -- a distinct file
    from the main budget.xlsb. Returns None if that file isn't present so the
    main dashboard build never breaks on it; the JRF tab is simply omitted
    when that happens (see build_html).

    Budget/Actual/Projection are all derived here rather than read from the
    Summary sheet's own (manually-typed) columns -- so month to month, the
    only upkeep needed on the sheet is the 'As of' date and the Annual -
    Actual (Cumulative) tab (which JRF was already updating monthly per that
    sheet's own instructions); everything else -- cumulative-to-date budget,
    variance, full-year projection, per category -- gets computed the same
    way the rest of this dashboard computes projections: actual-to-date plus
    whatever's still budgeted for the remaining months."""
    if not os.path.exists(FP_JRF):
        return None
    wb = openpyxl.load_workbook(FP_JRF, data_only=True)
    ws = wb['Summary']
    as_of_raw = ws.cell(row=4, column=2).value
    as_of_month, as_of_str = _jrf_as_of_month(as_of_raw)

    def norm_opex_cat(section, cat):
        return JRF_OPEX_CATEGORY_FIX.get(cat, cat) if section == 'opex' else cat

    # Budget: Monthly - Budget has Category/Account/Jan..Dec/Total columns.
    # Aggregated two ways -- by category (for the rev/opex category cards,
    # where category is the real distinguishing field) and by exact account
    # name (for Donations, whose category column is uniformly "Donation"
    # for all three lines -- the account name is what actually varies).
    budget_cum, budget_full, budget_rem = {}, {}, {}
    budget_cum_acct, budget_full_acct, budget_rem_acct = {}, {}, {}
    for section, cat, acct, months in _jrf_section_rows(
            wb['Monthly - Budget'], 0, 1, lambda r: [r[c] or 0 for c in range(2, 14)]):
        cat = norm_opex_cat(section, cat)
        cum, full, rem = sum(months[:as_of_month]), sum(months), sum(months[as_of_month:])
        budget_cum.setdefault(section, {})[cat]  = budget_cum.setdefault(section, {}).get(cat, 0) + cum
        budget_full.setdefault(section, {})[cat] = budget_full.setdefault(section, {}).get(cat, 0) + full
        budget_rem.setdefault(section, {})[cat]  = budget_rem.setdefault(section, {}).get(cat, 0) + rem
        budget_cum_acct.setdefault(section, {})[acct]  = cum
        budget_full_acct.setdefault(section, {})[acct] = full
        budget_rem_acct.setdefault(section, {})[acct]  = rem

    # Actual: Actual (Cumulative) - Sorted has Category/Account/2023../2026.
    actual_cum, actual_cum_acct = {}, {}
    for section, cat, acct, val in _jrf_section_rows(
            wb['Actual (Cumulative) - Sorted'], 0, 1, lambda r: r[5] or 0):
        cat = norm_opex_cat(section, cat)
        actual_cum.setdefault(section, {})[cat]  = actual_cum.setdefault(section, {}).get(cat, 0) + val
        actual_cum_acct.setdefault(section, {})[acct] = actual_cum_acct.setdefault(section, {}).get(acct, 0) + val

    def line(section, key, by_account=False):
        bc_src, bf_src, rem_src, ac_src = (
            (budget_cum_acct, budget_full_acct, budget_rem_acct, actual_cum_acct) if by_account
            else (budget_cum, budget_full, budget_rem, actual_cum)
        )
        bc  = bc_src.get(section, {}).get(key, 0)
        bf  = bf_src.get(section, {}).get(key, 0)
        rem = rem_src.get(section, {}).get(key, 0)
        ac  = ac_src.get(section, {}).get(key, 0)
        proj = ac + rem
        return dict(plan=bc, act=ac, var=ac-bc, ann_plan=bf, ann_proj=proj, ann_var=proj-bf)

    def total(section):
        bc  = sum(budget_cum.get(section, {}).values())
        ac  = sum(actual_cum.get(section, {}).values())
        bf  = sum(budget_full.get(section, {}).values())
        rem = sum(budget_rem.get(section, {}).values())
        proj = ac + rem
        return dict(plan=bc, act=ac, var=ac-bc, ann_plan=bf, ann_proj=proj, ann_var=proj-bf)

    jrf = {}
    jrf['as_of'] = 'As of ' + as_of_str
    jrf['rev'] = total('rev')
    jrf['rev_cats'] = [(c, line('rev', c)) for c in ['Corporate','DTC','Event','Other','Private','Raffle']]
    jrf['don'] = total('don')
    jrf['don_cats'] = [(c.replace('Donation - ',''), line('don', c, by_account=True)) for c in
        ['Donation - Community & Education','Donation - Environmental','Donation - Storm & Disaster Relief']]
    jrf['opex'] = total('opex')
    jrf['opex_cats'] = [(c, line('opex', c)) for c in ['COGS','Program Services','Event','Finance','Labor','MKG','Other']]
    # Net Income has no line of its own in Monthly - Budget / Actual
    # (Cumulative) - Sorted -- it's Revenue minus Donations minus OpEx,
    # same as the Summary sheet's own formula.
    jrf['net_income'] = {
        k: jrf['rev'][k] - jrf['don'][k] - jrf['opex'][k] for k in jrf['rev']
    }

    # Outstanding donation commitments (Name/Amount/Bucket), rows 16-30
    commitments = []
    for r in range(16, 31):
        name = ws.cell(row=r, column=10).value
        if name is None or str(name).strip() == 'Total': continue
        commitments.append((
            str(name).strip(),
            ws.cell(row=r, column=11).value or 0,
            ws.cell(row=r, column=12).value or '',
        ))
    jrf['commitments'] = commitments
    jrf['commitments_total'] = ws.cell(row=31, column=11).value or 0

    jrf['don_analysis'] = dict(
        donated_to_date=ws.cell(row=37, column=10).value or 0,
        projected=ws.cell(row=37, column=12).value or 0,
        committed=ws.cell(row=37, column=14).value or 0,
        available=ws.cell(row=37, column=15).value or 0,
    )

    # YoY totals, top-level only (Total Income / Total COGS / Total OpEx /
    # Net Income) -- prior years (2023-2025) only have a single annual
    # actual on file, not a monthly budget to compare against, so there's
    # no meaningful "cumulative to date" cut for past years the way there
    # is for the current year above; this chart stays at the clean,
    # unambiguous top-line full-year comparison instead.
    ws2 = wb['Actual (Cumulative) - Sorted']
    yoy = {}
    for row in ws2.iter_rows(min_row=7, max_row=ws2.max_row, values_only=True):
        acct = row[1]
        if acct is None: continue
        acct = str(acct).strip()
        if acct in ('Total Income', 'Total Cost of Goods Sold', 'Total Operating Expenses', 'Net Income'):
            yoy[acct] = [row[2] or 0, row[3] or 0, row[4] or 0, row[5] or 0]
    jrf['yoy'] = yoy
    jrf['yoy_years'] = ['2023', '2024', '2025', '2026 YTD']

    return jrf

# ── Design assets (fonts + component CSS extracted verbatim from the approved
#    mockups, so the live dashboard matches what was actually designed) ──────

_ASSETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'assets')

def _asset(name):
    with open(os.path.join(_ASSETS_DIR, name)) as f:
        return f.read()

FONTS_CSS = _asset('fonts.css')
RC_CSS    = _asset('rc_system.css')
LOGO_B64  = _asset('logo_base64.txt').strip().split(',', 1)[-1]
LABOR_ICONS_TEMPLATE = _asset('labor_icons_template.html')

def build_labor_fun_stats(pr):
    """The Employment Mix icon row (person/clock/$/gender/brand-ink icons)
    extracted verbatim from the approved mockup, with live values swapped in."""
    html = LABOR_ICONS_TEMPLATE
    for token, val in [
        ('__TOTAL_ACTIVE__', pr['total']),
        ('__FT__', pr['ft']), ('__PT__', pr['pt']),
        ('__SALARY__', pr['salary']), ('__HOURLY__', pr['hourly']),
        ('__MALE__', pr['male']), ('__FEMALE__', pr['female']),
        ('__BRAND__', pr['brand']), ('__INK__', pr['ink']),
    ]:
        html = html.replace(token, str(val))
    return html

EXTRA_CSS = '''
*{box-sizing:border-box}
.tab-panel{display:none}
.tab-panel.active{display:block}
.chart-wrap{position:relative;width:100%}
.rc-card{overflow-x:auto}
.rc-expand-btn{
  position:absolute;top:0;right:0;background:transparent;cursor:pointer;
  border:1px solid var(--surface-alt);border-radius:4px;padding:4px 10px;
  font-family:'IBM Plex Mono',monospace;font-size:11px;color:var(--ink);
}
.rc-expand-btn:hover{background:var(--surface)}
.chart-modal-overlay{
  display:none;position:fixed;inset:0;background:rgba(20,26,29,.6);
  z-index:1000;align-items:center;justify-content:center;padding:24px;
}
.chart-modal-overlay.open{display:flex}
.chart-modal-content{
  background:#fff;border-radius:8px;width:min(1400px,95vw);height:min(820px,90vh);
  display:flex;flex-direction:column;padding:20px 24px;
}
.chart-modal-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-shrink:0}
.chart-modal-title{font-family:var(--display);font-weight:700;font-size:18px;color:#252933}
.chart-modal-actions{display:flex;gap:8px;align-items:center}
.chart-modal-actions button{
  border:1px solid var(--surface-alt);border-radius:4px;background:#fff;cursor:pointer;
  padding:6px 14px;font-family:'IBM Plex Mono',monospace;font-size:12px;color:var(--ink);
}
.chart-modal-actions button:hover{background:var(--surface)}
.chart-modal-close{font-size:16px;line-height:1;padding:6px 10px !important}
.chart-modal-canvas-wrap{flex:1;position:relative;min-height:0}
'''

# ── rc- component helpers ────────────────────────────────────────────────────

def rc_stat(label, value, cls=''):
    c = ' ' + cls if cls else ''
    return ('<div class="rc-stat"><span class="rc-stat-label">' + label + '</span>'
            '<span class="rc-stat-val' + c + '">' + value + '</span></div>')

def rc_box(group_label, plan_val, act_lbl, act_val, var_val, var_cls):
    return ('<div class="rc-box"><div class="rc-group-label">' + group_label + '</div>'
            '<div class="rc-row">'
            + rc_stat('Plan', plan_val) + rc_stat(act_lbl, act_val) + rc_stat('Variance', var_val, var_cls)
            + '</div></div>')

def var_cls(v, favorable_if_below=False):
    good = (v <= 0) if favorable_if_below else (v >= 0)
    return 'pos' if good else 'neg'

def rc_card_kpi(title, cum, full, favorable_if_below=False, desc=None, body_extra='', trend=None):
    """cum/full = (plan, actual_or_proj, var). favorable_if_below flips the
    variance color convention for cost lines (Labor, OpEx-style spend).
    trend, if given, is a seasonality-adjusted full-year projection (see
    seasonal_ann_proj) shown as a third box alongside the plan-based one."""
    cum_plan, cum_act, cum_var = cum
    full_plan, full_proj, full_var = full
    trend_box = ''
    if trend is not None:
        trend_var = trend - full_plan
        trend_box = rc_box('Trend (Seasonality-Adj.)', fk(full_plan), 'Trend Proj.', fk(trend),
                            vk(trend_var), var_cls(trend_var, favorable_if_below))
    return (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-headrow"><div class="rc-name">' + title + '</div>'
        + ('<div class="rc-desc">' + desc + '</div>' if desc else '') +
        '</div>'
        '<div class="rc-boxrow">'
        + rc_box('Cumulative to Date', fk(cum_plan), 'Actual', fk(cum_act), vk(cum_var), var_cls(cum_var, favorable_if_below))
        + rc_box('Full Year', fk(full_plan), 'Projection', fk(full_proj), vk(full_var), var_cls(full_var, favorable_if_below))
        + trend_box +
        '</div>'
        + body_extra +
        '</div>\n'
    )

def rc_trend_chart(chart_id, title, desc=None):
    """A 52-week canvas placed under a summary card -- see
    build_summary_trend_charts_js() for the four-line (2025 Actual, 2026
    Plan, 2026 Actual, 2026 Trend) series it renders into this canvas. The
    expand button opens the same chart full-screen via a delegated click
    listener on .rc-expand-btn (see CHART_MODAL_JS) with a PNG download
    option -- data attributes rather than an inline onclick, since the
    title's own double quotes (from embedding it as a JS string) would
    otherwise collide with the HTML attribute's double quotes."""
    return (
        '<div class="rc-card" style="grid-column:1 / -1;position:relative">'
        '<div class="rc-headrow"><div class="rc-name">' + title + '</div>'
        + ('<div class="rc-desc">' + desc + '</div>' if desc else '') +
        '<button class="rc-expand-btn" data-chart-id="' + chart_id + '" data-chart-title="' + html_escape(title) + '">⤢ Expand</button>'
        '</div>'
        '<div style="height:220px"><canvas id="' + chart_id + '"></canvas></div>'
        '</div>\n'
    )

def rc_divider(title):
    return '<div class="rc-divider"><div class="rc-divider-title">' + title + '</div></div>\n'

def rc_hltable(rows, is_cost=False):
    """rows: list of (label, ytd_plan, ytd_act, ann_plan, ann_proj, highlighted).
    Matches the approved design's 5-metric itemized table (Account Line / YTD
    Plan / YTD Actual / YTD Var / Ann. Proj. / Ann. Var — no Ann. Plan column,
    since under the actual+remaining-plan projection model Ann. Var always
    equals YTD Var, so showing both variance columns is the useful pair)."""
    body = ''
    for name, ytd_plan, ytd_act, ann_plan, ann_proj, hl in rows:
        ytd_var = ytd_act - ytd_plan
        ann_var = ann_proj - ann_plan
        star = '<span class="rc-hl-star">&#9733;</span>' if hl else ''
        body += (
            '<tr>'
            '<td>' + star + name + '</td>'
            '<td>' + fk(ytd_plan) + '</td>'
            '<td>' + fk(ytd_act) + '</td>'
            '<td class="' + var_cls(ytd_var, is_cost) + '">' + vk(ytd_var) + '</td>'
            '<td>' + fk(ann_proj) + '</td>'
            '<td class="' + var_cls(ann_var, is_cost) + '">' + vk(ann_var) + '</td>'
            '</tr>\n'
        )
    return (
        '<div class="rc-hl">'
        '<table class="rc-hltable">'
        '<colgroup><col class="rc-hl-col-name">'
        '<col class="rc-hl-col-stat"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat">'
        '<col class="rc-hl-col-stat"><col class="rc-hl-col-stat"></colgroup>'
        '<thead><tr><th>Account Line</th><th>YTD Plan</th><th>YTD Actual</th><th>YTD Var</th>'
        '<th>Ann. Proj.</th><th>Ann. Var</th></tr></thead>'
        '<tbody>' + body + '</tbody>'
        '</table></div>'
    )

def ch_card(name, ytd_act, ytd_plan, ann_plan, ann_proj, trend_proj=None):
    """trend_proj is optional -- omit it for lines with no seasonality curve
    yet (e.g. JRF's category breakdown) and the trend footer is skipped."""
    var = ytd_act - ytd_plan
    cls = var_cls(var)
    pct = max(0, min(100, (ytd_act / ann_proj * 100) if ann_proj else 0))
    trend_row = ''
    if trend_proj is not None:
        trend_var = trend_proj - ann_plan
        trend_cls = var_cls(trend_var)
        trend_row = (
            '<div class="ch-foot ch-trend ' + trend_cls + '">Trend Proj. ' + fk(trend_proj)
            + ' (' + vk(trend_var) + ' vs plan)</div>'
        )
    return (
        '<div class="ch-card">'
        '<div class="ch-name">' + name + '</div>'
        '<div class="ch-value">' + fk(ytd_act) + '</div>'
        '<div class="ch-delta ' + cls + '">' + vk(var) + ' vs plan</div>'
        '<div class="ch-bar"><div class="ch-bar-fill ' + cls + '" style="width:' + f'{pct:.0f}' + '%"></div></div>'
        '<div class="ch-foot">Ann. Proj. ' + fk(ann_proj) + '</div>'
        + trend_row +
        '</div>'
    )

def dept_tile(name, value):
    return (
        '<div class="dept-tile">'
        '<div class="dept-tile-top"><span class="dept-tile-name">' + name + '</span></div>'
        '<div class="dept-tile-val">' + value + '</div>'
        '</div>'
    )

def rc_bignum_card(title, value, sub, sub2=None, sub2_cls='pos'):
    return (
        '<div class="rc-card">'
        '<div class="rc-group-label">' + title + '</div>'
        '<div class="rc-stat-val" style="font-size:26px;margin:6px 0 4px">' + value + '</div>'
        '<div class="rc-desc">' + sub + '</div>'
        + (('<div class="rc-desc ' + sub2_cls + '" style="font-family:var(--mono);margin-top:2px">' + sub2 + '</div>') if sub2 else '')
        + '</div>'
    )

# ── Chart JS ─────────────────────────────────────────────────────────────────

def build_chart_js(d):
    cf     = d['cf']
    labels = '[' + ','.join('"' + l + '"' for l in cf['labels']) + ']'
    cout_tot = [b+i+o+l+e for b,i,o,l,e in zip(cf['cout_b'],cf['cout_i'],cf['cout_o'],cf['cout_l'],cf['cout_e'])]

    return (
        'Chart.defaults.font.family="\'IBM Plex Mono\',monospace";Chart.defaults.font.size=10;\n'
        'const fmtK=v=>{const a=Math.abs(v);'
        'if(a>=1000000)return"$"+(v<0?"−":"")+(a/1000000).toFixed(2)+"M";'
        'if(a>=1000)return"$"+(v<0?"−":"")+(a/1000).toFixed(1)+"K";'
        'return"$"+v.toLocaleString();};\n'
        # Estimated points (a week with no recorded reading, carried forward
        # from the prior week rather than dropped to zero) render as hollow
        # markers instead of solid, and say so in their tooltip.
        'const estNote=c=>(c.raw&&c.raw.e)?" (not recorded — carried fwd)":"";\n'
        'const estStyle=color=>({'
        'pointBackgroundColor:ctx=>(ctx.raw&&ctx.raw.e)?"#fff":color,'
        'pointBorderColor:color,'
        'pointBorderWidth:ctx=>(ctx.raw&&ctx.raw.e)?2:1,'
        'pointRadius:ctx=>(ctx.raw&&ctx.raw.e)?4:3,'
        '});\n'
        'const WKS=' + labels + ';\n'
        '(function(){const el=document.getElementById("cfChart");if(!el)return;'
        'new Chart(el,{type:"line",data:{labels:WKS,datasets:['
        '{label:"Cash in", data:' + js_arr(cf['cin_t'])  + ',borderColor:"#2F7A5C",backgroundColor:"transparent",borderWidth:2.5,pointRadius:3,pointHoverRadius:5,tension:0.3},'
        '{label:"Cash out",data:' + js_arr(cout_tot)     + ',borderColor:"#B4482F",backgroundColor:"transparent",borderWidth:2.5,pointRadius:3,pointHoverRadius:5,tension:0.3},'
        ']},options:{responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:true,position:"top"}},'
        'scales:{x:{grid:{color:"#EDECED"}},y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}'
        '}});})();\n'
        'const HL=Array.from({length:52},(_,i)=>"Wk "+(i+1));\n'
        'const histOpts={responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmtK(c.parsed.y)+estNote(c)}}},'
        'scales:{x:{type:"category",grid:{color:"#EDECED"},ticks:{maxRotation:45,callback:function(v,i){return i%4===0?this.getLabelForValue(v):"";}}},y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}};\n'
        '(function(){const el=document.getElementById("invHistCOG");if(!el)return;'
        'new Chart(el,{type:"line",data:{labels:HL,datasets:['
        '{label:"2024",data:' + to_xy(d['cog_2024'], d['inv_est']['cog_2024'])   + ',borderColor:"#B0B4B8",backgroundColor:"transparent",borderWidth:2,tension:0.3,parsing:false,...estStyle("#B0B4B8")},'
        '{label:"2025",data:' + to_xy(d['cog_2025'], d['inv_est']['cog_2025'])   + ',borderColor:"#586D72",backgroundColor:"transparent",borderWidth:2,tension:0.3,parsing:false,...estStyle("#586D72")},'
        '{label:"2026",data:' + to_xy(d['cog_2026'], d['inv_est']['cog_2026'])  + ',borderColor:"#43575E",backgroundColor:"transparent",borderWidth:2.5,tension:0.3,parsing:false,...estStyle("#43575E")},'
        ']},options:histOpts});})();\n'
        '(function(){const el=document.getElementById("invTotUnits");if(!el)return;'
        'new Chart(el,{type:"line",data:{labels:HL,datasets:['
        '{label:"2024",data:' + to_xy(d['units_2024'], d['inv_est']['units_2024']) + ',borderColor:"#B0B4B8",backgroundColor:"transparent",borderWidth:2,tension:0.3,parsing:false,...estStyle("#B0B4B8")},'
        '{label:"2025",data:' + to_xy(d['units_2025'], d['inv_est']['units_2025']) + ',borderColor:"#586D72",backgroundColor:"transparent",borderWidth:2,tension:0.3,parsing:false,...estStyle("#586D72")},'
        '{label:"2026",data:' + to_xy(d['units_2026'], d['inv_est']['units_2026'])+ ',borderColor:"#43575E",backgroundColor:"transparent",borderWidth:2.5,tension:0.3,parsing:false,...estStyle("#43575E")},'
        ']},options:histOpts});})();\n'
        'const stOpts={responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.dataset.label+": "+c.parsed.y.toFixed(1)+"%"+estNote(c)}}},'
        'scales:{x:{type:"category",grid:{color:"#EDECED"},ticks:{maxRotation:45,callback:function(v,i){return i%4===0?this.getLabelForValue(v):"";}}},y:{min:0,max:100,grid:{color:"#EDECED"},ticks:{callback:v=>v+"%"}}}};\n'
        'const SL=Array.from({length:52},(_,i)=>"Wk "+(i+1));\n'
        'function mkST(id,ds){const el=document.getElementById(id);if(!el)return;'
        'new Chart(el,{type:"line",data:{labels:SL,datasets:ds.map(d=>{return{label:d.l,data:d.d,'
        'borderColor:d.c,backgroundColor:"transparent",borderWidth:2,tension:0.3,parsing:false,...estStyle(d.c)};})},'
        'options:stOpts});}\n'
        'mkST("invSSSellThru",['
        '{l:"S24",d:' + st_xy(d['st_s24'], d['inv_est']['st_s24']) + ',c:"#B0B4B8"},'
        '{l:"S25",d:' + st_xy(d['st_s25'], d['inv_est']['st_s25']) + ',c:"#586D72"},'
        '{l:"S26",d:' + st_xy(d['st_s26'], d['inv_est']['st_s26']) + ',c:"#43575E"}]);\n'
        'mkST("invFWSellThru",['
        '{l:"F24",d:' + st_xy(d['st_f24'], d['inv_est']['st_f24']) + ',c:"#B0B4B8"},'
        '{l:"F25",d:' + st_xy(d['st_f25'], d['inv_est']['st_f25']) + ',c:"#586D72"},'
        '{l:"F26",d:' + st_xy(d['st_f26'], d['inv_est']['st_f26']) + ',c:"#43575E"}]);\n'
        + build_labor_chart_js(d)
        + build_summary_trend_charts_js(d)
        + build_cf_ty_ly_charts_js(d)
        + build_ar_chart_js(d)
        + (build_jrf_chart_js(d['jrf']) if d.get('jrf') else '')
    )

def build_summary_trend_charts_js(d):
    """52-week line chart for each Summary card: 2025 Actual (reference),
    2026 Plan, 2026 Actual (solid, stops at the current week), and 2026
    Trend (dashed continuation from the current week to week 52) -- see
    the trend_charts computation in read_all()."""
    tc = d['trend_charts']
    sections = [
        ('chart-trend-revenue',    'revenue'),
        ('chart-trend-cogs',       'cogs'),
        ('chart-trend-opex',       'opex'),
        ('chart-trend-net_income', 'net_income'),
    ]
    # Options are built by a factory function, called fresh for every chart
    # instance (including the modal's) -- Chart.js writes internal resolver
    # state onto whatever options object it's given, so reusing one live
    # instance's already-resolved .options for a second instance breaks it.
    # data.datasets, by contrast, is plain and safe to share by reference
    # (this is Chart.js's own supported pattern for one dataset in multiple
    # charts), so window.__chartData is shared as-is with the modal.
    out = (
        'function sumTrendOpts(){return {responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:true,position:"top"},tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmtK(c.parsed.y)}}},'
        'scales:{x:{type:"category",grid:{color:"#EDECED"},ticks:{maxRotation:45,callback:function(v,i){return i%4===0?this.getLabelForValue(v):"";}}},'
        'y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}};}\n'
        'window.__chartData=window.__chartData||{};window.__chartOptsFn=window.__chartOptsFn||{};window.__charts=window.__charts||{};\n'
    )
    for chart_id, key in sections:
        s = tc[key]
        out += (
            '(function(){const el=document.getElementById("' + chart_id + '");if(!el)return;'
            'const data={labels:HL,datasets:['
            '{label:"2025 Actual",data:' + trend_xy(s['act_2025'])   + ',borderColor:"#B0B4B8",backgroundColor:"transparent",borderWidth:2,pointRadius:2,tension:0.3,parsing:false},'
            '{label:"2026 Plan",data:'   + trend_xy(s['plan_2026'])  + ',borderColor:"#586D72",backgroundColor:"transparent",borderWidth:1.5,borderDash:[4,3],pointRadius:0,tension:0.3,parsing:false},'
            '{label:"2026 Actual",data:' + trend_xy(s['act_2026'])   + ',borderColor:"#43575E",backgroundColor:"transparent",borderWidth:2.5,pointRadius:2,tension:0.3,parsing:false},'
            '{label:"2026 Trend",data:'  + trend_xy(s['trend_2026']) + ',borderColor:"#43575E",backgroundColor:"transparent",borderWidth:2.5,borderDash:[6,4],pointRadius:0,tension:0.3,parsing:false},'
            ']};'
            'window.__chartData["' + chart_id + '"]=data;window.__chartOptsFn["' + chart_id + '"]=sumTrendOpts;'
            'window.__charts["' + chart_id + '"]=new Chart(el,{type:"line",data:data,options:sumTrendOpts()});})();\n'
        )
    return out

CF_TY_LY_COLORS = dict(
    bal_ty="#43575E", bal_ly="#B0B4B8",     # slate: Total Bank Balance, TY solid / LY light
    in_ty="#2F7A5C", in_ly="#A8D5C2",       # green: Cash In, TY solid / LY light
    out_ty="#B4482F", out_ly="#E8B7A8",     # red: Cash Out, TY solid / LY light
    draw_ty="#2B6CB0", draw_ly="#9DC3E6",   # blue: CL Draw, TY solid / LY light
    pay_ty="#7B3F9E", pay_ly="#C9AEDB",     # purple: CL Paydown, TY solid / LY light
)

def build_cf_ty_ly_charts_js(d):
    """Section #1 of the rebuilt Cash Flow tab, this year vs. last:
    - Total Bank Balance (lines): the week's starting balance, TY vs. LY.
    - Weekly (bars): Cash In/Cash Out only, TY vs. LY.
    - CL Draw / Paydown (lines, directly below Weekly): net Columbia CL
      activity per week (draw minus paydown), TY vs. LY -- one line per
      year rather than 4 separate bars, since Draw/Paydown are baked into
      the raw Cash In/Out totals above (a draw is a real "Receive Money"
      transaction, a paydown a real debit) and this chart exists purely
      as the callout, not a second full breakdown: up = net borrowing
      that week, down = net paydown.
    All three read from cfw_history (see read_cash_flow_weekly_history) --
    returns '' if that's empty (sheet missing/unreadable)."""
    hist = d.get('cfw_history') or {}
    ty_year = max(hist.keys()) if hist else None
    ly_year = ty_year - 1 if ty_year else None
    ty = hist.get(ty_year, {})
    ly = hist.get(ly_year, {})
    if not ty:
        return ''

    def get(block, key):
        return block.get(key) or [None] * 52

    def net(draw, pay):
        """Net CL activity per week (draw minus paydown), zero-filled across
        all 52 weeks. The sheet only carries a figure in weeks where a draw
        or paydown actually happened -- 19 of 52 weeks in 2025, 3 in 2026 --
        so treating the blanks as gaps left the chart as disconnected dots
        and stray fragments. A blank week here genuinely means "no draw, no
        paydown," i.e. zero activity, not unknown, so both years plot as one
        continuous line across the full year."""
        return [round((dv or 0) - (pv or 0), 2) for dv, pv in zip(draw, pay)]

    bal_ty, bal_ly = get(ty, 'bank_balance'), get(ly, 'bank_balance')
    in_ty, out_ty = get(ty, 'cash_in'), [(-v if v is not None else None) for v in get(ty, 'cash_out')]
    in_ly, out_ly = get(ly, 'cash_in'), [(-v if v is not None else None) for v in get(ly, 'cash_out')]
    cl_ty = net(get(ty, 'cl_draw'), get(ty, 'cl_paydown'))
    cl_ly = net(get(ly, 'cl_draw'), get(ly, 'cl_paydown'))

    labels = '[' + ','.join('"Wk ' + str(i) + '"' for i in range(1, 53)) + ']'
    c = CF_TY_LY_COLORS

    def bar_ds(label, arr, color):
        return ('{type:"bar",label:"' + label + '",data:' + js_arr_n(arr) +
                ',backgroundColor:"' + color + '",borderRadius:2,order:2}')
    def line_ds(label, arr, color, dashed):
        dash = ',borderDash:[5,3]' if dashed else ''
        return ('{type:"line",label:"' + label + '",data:' + js_arr_n(arr) +
                ',borderColor:"' + color + '",backgroundColor:"transparent",'
                'borderWidth:1.75,pointRadius:1.5,tension:0.2,order:1' + dash + '}')

    balance_datasets = ','.join([
        line_ds('Bank Balance (TY)', bal_ty, c['bal_ty'], False),
        line_ds('Bank Balance (LY)', bal_ly, c['bal_ly'], True),
    ])

    weekly_datasets = ','.join([
        bar_ds('Cash In (TY)', in_ty, c['in_ty']), bar_ds('Cash In (LY)', in_ly, c['in_ly']),
        bar_ds('Cash Out (TY)', out_ty, c['out_ty']), bar_ds('Cash Out (LY)', out_ly, c['out_ly']),
    ])

    cl_datasets = ','.join([
        line_ds('Columbia CL (TY)', cl_ty, c['draw_ty'], False),
        line_ds('Columbia CL (LY)', cl_ly, c['draw_ly'], False),
    ])

    return (
        'function cfTyLyOpts(){return {responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:true,position:"top",labels:{boxWidth:12,font:{size:9}}},'
        'tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmtK(c.parsed.y)}}},'
        'scales:{x:{type:"category",grid:{display:false},'
        'ticks:{maxRotation:45,callback:function(v,i){return i%4===0?this.getLabelForValue(v):"";}}},'
        'y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}};}\n'
        'window.__chartData=window.__chartData||{};window.__chartOptsFn=window.__chartOptsFn||{};window.__charts=window.__charts||{};\n'
        '(function(){const el=document.getElementById("chart-cf-ty-ly-balance");if(!el)return;'
        'const data={labels:' + labels + ',datasets:[' + balance_datasets + ']};'
        'window.__chartData["chart-cf-ty-ly-balance"]=data;window.__chartOptsFn["chart-cf-ty-ly-balance"]=cfTyLyOpts;'
        'window.__charts["chart-cf-ty-ly-balance"]=new Chart(el,{type:"line",data:data,options:cfTyLyOpts()});})();\n'
        '(function(){const el=document.getElementById("chart-cf-ty-ly-weekly");if(!el)return;'
        'const data={labels:' + labels + ',datasets:[' + weekly_datasets + ']};'
        'window.__chartData["chart-cf-ty-ly-weekly"]=data;window.__chartOptsFn["chart-cf-ty-ly-weekly"]=cfTyLyOpts;'
        'window.__charts["chart-cf-ty-ly-weekly"]=new Chart(el,{type:"bar",data:data,options:cfTyLyOpts()});})();\n'
        '(function(){const el=document.getElementById("chart-cf-ty-ly-cl-weekly");if(!el)return;'
        'const data={labels:' + labels + ',datasets:[' + cl_datasets + ']};'
        'window.__chartData["chart-cf-ty-ly-cl-weekly"]=data;window.__chartOptsFn["chart-cf-ty-ly-cl-weekly"]=cfTyLyOpts;'
        'window.__charts["chart-cf-ty-ly-cl-weekly"]=new Chart(el,{type:"line",data:data,options:cfTyLyOpts()});})();\n'
    )

def build_summary_panel(d):
    """Top-of-page Summary panel, sitting above every other Cash Flow tab
    section -- a compact EOY cash projection, built from Jeremy's own
    Cash_Flow_V2 outline (2026-09-02). Returns '' if the current week's
    Bank Balance isn't available yet (nothing to project from).

    Cash In side: current Bank Balance + last year's actual Cash In for
    the weeks remaining this year (a conservative, actuals-based
    projection rather than the 2026 budget's assumed pace).

    Cash Out side: 2026 Budget remaining-weeks Labor + OpEx (unchanged
    from Cash Bridge's own formulas), 2026 Budget remaining-weeks COGS
    *excluding* the four lines now tracked live via A/P (Spring/Summer,
    Fall/Winter, Special, Screen Printing Wholesale -- same exclusion
    Cash Bridge already applies), plus the live A/P - Key Vendors and
    A/P - CNS balances themselves, taken whole -- no aging filter yet on
    Key Vendors (Jeremy: "we can set some rules based on the aging in AT
    later"), full outstanding balance on CNS by design (so he can weigh
    paying vendors down vs. drawing the credit line with the complete
    picture).

    Both A/P figures come from the new 'AP - Key' / 'AP - CNS' sheets --
    hand-entered cumulative snapshots, not backfilled week by week, so
    latest_nonzero() takes the most recently entered week rather than
    assuming the current week itself is populated.

    Below Available Cash: an 'A/P - Key Vendors (Not Yet Due)' provision,
    added back per Jeremy's aging rules (2026-09-03) -- he only actually
    needs to pay what's 31+ days late, so the Current + 1-30 Days portion
    of the full balance above is cash that doesn't really need to go out
    yet. S&S Activewear is an exception (2026-09-03): only its 90+ Days
    bucket is a must-pay, so its Current/1-30/31-60/61-90 all count as
    not-yet-due. Sourced live from ap_key_vendors_open() (WHSL PMT
    Tracker due dates vs. today), independent of the AP - Key XL tab's
    hand-entered total used above.

    Last three rows (2026-09-03): a static Columbia Bank Credit Line (EOY
    Goal) of $1.7M; A/P - CNS (Carryover debt to 2027) -- the shortfall,
    floored at 0, between what it'd take to pay the credit line down to
    that goal (Current CL Balance − EOY Goal) and Available Cash
    (Adjusted); and Cash Surplus (2027 BOY) -- the complementary
    leftover, also floored at 0, once that paydown is fully funded.
    Exactly one of the two is normally nonzero. The idea: if there isn't
    enough cash to both fund the paydown and pay CNS in full, the unpaid
    CNS balance is what rolls to next year rather than drawing the
    credit line further; if there's more than enough, what's left over
    is real cash carried into 2027.

    Two rows from Jeremy's outline are intentionally omitted: Projected
    Credit Line Paydown and A/P CNS Carryforward (2027) -- both still TBD
    on his end, pending how he wants to model vendor-pay-vs-credit-line-
    draw tradeoffs."""
    hist = d.get('cfw_history') or {}
    ty_year = max(hist.keys()) if hist else None
    ly_year = ty_year - 1 if ty_year else None
    if not ty_year:
        return ''
    WK = d['week']

    def get(yr, key):
        return (hist.get(yr) or {}).get(key) or [None] * 52

    bank_bal_arr = get(ty_year, 'bank_balance')
    starting_bank_balance = bank_bal_arr[WK - 1] if WK <= len(bank_bal_arr) else None
    if starting_bank_balance is None:
        return ''

    ly_cash_in = get(ly_year, 'cash_in')
    cash_in_projected = sum(v for v in ly_cash_in[WK:] if v is not None)
    cash_in_eoy = starting_bank_balance + cash_in_projected

    rem_labor_plan = d['labor_total']['ann_plan'] - d['labor_total']['ytd_plan']
    rem_opex_plan  = d['opex']['ann_plan'] - d['opex']['plan']
    other_cogs_out = (d['cogs']['ann_plan'] - d['cogs']['plan']) - d['cogs_ap_remaining_plan']

    ap_key_total = latest_nonzero((d.get('ap_key_xl') or {}).get(ty_year)) or 0.0
    ap_cns_total = latest_nonzero((d.get('ap_cns_xl') or {}).get(ty_year)) or 0.0

    cash_out_eoy = rem_labor_plan + other_cogs_out + ap_key_total + ap_cns_total + rem_opex_plan
    available_cash = cash_in_eoy - cash_out_eoy

    cl_balance_arr = get(ty_year, 'cl_balance')
    cl_balance = cl_balance_arr[WK - 1] if WK <= len(cl_balance_arr) else None

    def row(label, value, note='', bold=False, top=False):
        weight = 'font-weight:700' if bold else ''
        topb = 'border-top:1.5px solid var(--ink)' if top else ''
        return (
            '<tr style="' + topb + '">'
            '<td style="' + weight + '">' + label + '</td>'
            '<td style="' + weight + '">' + value + '</td>'
            '<td style="text-align:left;font-family:var(--body);font-weight:400;opacity:.75;padding-left:16px">' + note + '</td>'
            '</tr>\n'
        )
    def spacer_row():
        return '<tr><td colspan="3" style="height:14px;border:none"></td></tr>\n'

    rows = row('Week', str(WK))
    rows += row('Starting Bank Balance', fk(starting_bank_balance),
                'Bank Balance sheet, week ' + str(WK) + ' Total.', top=True)
    rows += row('Total Cash IN (projected)', fk(cash_in_projected),
                str(ly_year) + "'s actual Cash In for weeks " + str(WK + 1) + '-52 '
                '(a conservative estimate for the rest of the year).')
    rows += row('Total Cash IN Projection (EOY)', fk(cash_in_eoy),
                'Starting Bank Balance + Total Cash IN (projected).', bold=True, top=True)

    rows += spacer_row()
    rows += row('Labor Out', fk(rem_labor_plan), '2026 Budget remaining weeks.', top=True)
    rows += row('Other COGS Out', fk(other_cogs_out),
                '2026 Budget remaining weeks, excluding Spring/Summer, Fall/Winter, Special, and Screen '
                'Printing Wholesale -- those are covered by A/P below.')
    rows += row('A/P - Key Vendors', fk(ap_key_total),
                'AP - Key tab, latest entered week TOTAL -- full balance, no aging filter yet.')
    rows += row('A/P - CNS', fk(ap_cns_total),
                'AP - CNS tab, latest entered week TOTAL -- full balance.')
    rows += row('OpEx', fk(rem_opex_plan), '2026 Budget remaining weeks.')
    rows += row('Total Cash OUT Projection (EOY)', fk(cash_out_eoy),
                'Labor Out + Other COGS Out + A/P - Key Vendors + A/P - CNS + OpEx.', bold=True, top=True)

    rows += spacer_row()
    rows += row('Available Cash', fk(available_cash),
                'Total Cash IN Projection (EOY) − Total Cash OUT Projection (EOY).', bold=True, top=True)

    # A/P - Key Vendors above is the full outstanding balance -- in practice
    # Jeremy only needs to pay down what's actually overdue past each
    # vendor's own threshold, so the not-yet-due portion (still within/just
    # past terms) is cash that doesn't really need to go out yet. Added back
    # here as its own line per his instruction, rather than silently
    # trimming the A/P - Key Vendors Cash Out line above. Sourced from live
    # per-PO aging (WHSL PMT Tracker due dates vs. today), not the AP - Key
    # XL tab's hand-entered total -- the two run close (within ~1% as of
    # this build) but aren't the same figure.
    #
    # S&S Activewear is a standing exception: Jeremy only needs to pay S&S
    # what's 90+ days overdue, not 30+ like every other Key Vendor -- so
    # S&S's 31-60 and 61-90 buckets count as "not yet due" too, on top of
    # Current + 1-30. Not-yet-due is just everything that isn't a must-pay
    # bucket for that vendor (see kv_must_pay_buckets), so this and the
    # Open POs by Vendor table's red columns can't drift apart.
    ALL_BUCKETS = ('current', 'd1_30', 'd31_60', 'd61_90', 'd90_plus')
    def kv_not_yet_due(v):
        must_pay = kv_must_pay_buckets(v['vendor'])
        return sum(v[b] for b in ALL_BUCKETS if b not in must_pay)
    kv = d.get('ap_key_vendors_open')
    if kv:
        kv_under_30 = sum(kv_not_yet_due(v) for v in kv['vendors'])
        available_cash_adj = available_cash + kv_under_30
        rows += row('A/P - Key Vendors (Not Yet Due)', vk(kv_under_30),
                     'Current + 1-30 Days for every vendor except S&S Activewear -- S&S only needs its '
                     '90+ Days bucket paid, so its Current/1-30/31-60/61-90 buckets all count here. From '
                     'WHSL PMT Tracker’s live per-PO due dates.')
        rows += row('Available Cash (Adjusted)', fk(available_cash_adj),
                     'Available Cash + A/P - Key Vendors (Not Yet Due) -- S&S Activewear excluded up to '
                     '90 days overdue (only its 90+ Days bucket is a must-pay); every other Key Vendor '
                     'excluded up to 30 days overdue (only 31+ days is a must-pay).', bold=True, top=True)

    if cl_balance is not None:
        rows += row('Columbia Credit Line Balance (Current)', fk(cl_balance),
                     'Cash Flow - Weekly, week ' + str(WK) + ' Columbia CL Balance.', top=True)

        cl_eoy_goal = 1_700_000.0
        rows += row('Columbia Bank Credit Line (EOY Goal)', fk(cl_eoy_goal),
                     'Target Columbia CL balance by year-end.')

        base_cash = available_cash_adj if kv else available_cash
        base_cash_label = 'Available Cash (Adjusted)' if kv else 'Available Cash'
        cl_paydown_needed = max(0.0, cl_balance - cl_eoy_goal)
        cns_carryover = max(0.0, cl_paydown_needed - base_cash)
        rows += row('A/P - CNS (Carryover debt to 2027)', fk(cns_carryover),
                     'Columbia CL paydown needed to reach the EOY Goal (Current − EOY Goal) minus '
                     + base_cash_label + ' -- the shortfall, if any, once that paydown is funded is A/P - '
                     'CNS left unpaid this year, carried to 2027.', top=True)

        cash_surplus_2027 = max(0.0, base_cash - cl_paydown_needed)
        rows += row('Cash Surplus (2027 BOY)', fk(cash_surplus_2027),
                     base_cash_label + ' left over after fully funding the Columbia CL paydown to the EOY '
                     'Goal -- $0 whenever there isn\'t enough to fully fund it (see A/P - CNS Carryover '
                     'above instead). The cash actually sitting in the bank at the start of 2027.')

    return (
        rc_divider('Summary — Week ' + str(WK))
        + '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-hl"><table class="rc-hltable">'
        '<colgroup><col style="width:32%"><col style="width:18%"><col style="width:50%"></colgroup>'
        '<thead><tr><th>Line</th><th>Value</th><th style="text-align:left;padding-left:16px">Notes</th></tr></thead>'
        '<tbody>' + rows + '</tbody>'
        '</table></div>'
        '</div>\n'
    )

def build_cf_ty_ly_section(d):
    """HTML for Section #1 of the rebuilt Cash Flow tab -- see
    build_cf_ty_ly_charts_js() for the data/series this renders.
    Returns '' if cfw_history has no data for the current year."""
    hist = d.get('cfw_history') or {}
    ty_year = max(hist.keys()) if hist else None
    if not ty_year or not hist.get(ty_year):
        return ''
    ly_year = ty_year - 1

    def card(chart_id, title, desc):
        return (
            '<div class="rc-card" style="grid-column:1 / -1;position:relative">'
            '<div class="rc-headrow"><div class="rc-name">' + title + '</div>'
            '<div class="rc-desc">' + desc + '</div>'
            '<button class="rc-expand-btn" data-chart-id="' + chart_id + '" data-chart-title="' + html_escape(title) + '">⤢ Expand</button>'
            '</div>'
            '<div style="height:320px;margin-top:10px"><canvas id="' + chart_id + '"></canvas></div>'
            '</div>\n'
        )

    balance_desc = 'Starting bank balance by week.'
    weekly_desc = 'Cash In vs. Cash Out by week.'
    cl_desc = ('Net Columbia CL activity by week (draw minus paydown) -- already included in '
               'Cash In/Cash Out above, called out separately since a draw week can otherwise '
               'read as unusually strong collections, or a paydown week as unusually heavy spend.')
    return (
        rc_divider("Cash In / Cash Out — " + str(ty_year) + " vs. " + str(ly_year))
        + card('chart-cf-ty-ly-balance', 'Total Bank Balance', balance_desc)
        + card('chart-cf-ty-ly-weekly', 'Weekly', weekly_desc)
        + card('chart-cf-ty-ly-cl-weekly', 'Columbia CL Draw / Paydown', cl_desc)
    )

AR_COLORS = dict(brand="#2B6CB0", ink="#B4482F")  # blue: Brand A/R, red: INK A/R

def build_ar_chart_js(d):
    """Total A/R trend by week, split Brand vs. INK -- from the AR sheet
    (see read_ar_history). 2026 only, no prior-year block exists on that
    sheet to compare against. Returns '' if ar_history is empty."""
    hist = d.get('ar_history') or {}
    year = max(hist.keys()) if hist else None
    if not year:
        return ''
    yr_data = hist[year]
    brand = yr_data.get('brand_total') or [None] * 52
    ink = yr_data.get('ink_total') or [None] * 52
    labels = '[' + ','.join('"Wk ' + str(i) + '"' for i in range(1, 53)) + ']'
    c = AR_COLORS

    def line_ds(label, arr, color):
        return ('{label:"' + label + '",data:' + js_arr_n(arr) +
                ',borderColor:"' + color + '",backgroundColor:"transparent",'
                'borderWidth:2,pointRadius:2,tension:0.25}')

    datasets = ','.join([line_ds('Brand A/R', brand, c['brand']), line_ds('INK A/R', ink, c['ink'])])

    return (
        'function arChartOpts(){return {responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:true,position:"top",labels:{boxWidth:12,font:{size:9}}},'
        'tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmtK(c.parsed.y)}}},'
        'scales:{x:{type:"category",grid:{display:false},'
        'ticks:{maxRotation:45,callback:function(v,i){return i%4===0?this.getLabelForValue(v):"";}}},'
        'y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}};}\n'
        'window.__chartData=window.__chartData||{};window.__chartOptsFn=window.__chartOptsFn||{};window.__charts=window.__charts||{};\n'
        '(function(){const el=document.getElementById("chart-ar-weekly");if(!el)return;'
        'const data={labels:' + labels + ',datasets:[' + datasets + ']};'
        'window.__chartData["chart-ar-weekly"]=data;window.__chartOptsFn["chart-ar-weekly"]=arChartOpts;'
        'window.__charts["chart-ar-weekly"]=new Chart(el,{type:"line",data:data,options:arChartOpts()});})();\n'
    )

def build_ar_section(d):
    """HTML for the A/R trend chart -- see build_ar_chart_js() for the
    data/series this renders. Returns '' if ar_history has no data."""
    hist = d.get('ar_history') or {}
    year = max(hist.keys()) if hist else None
    if not year or not hist.get(year):
        return ''
    return (
        rc_divider("Accounts Receivable — " + str(year))
        + '<div class="rc-card" style="grid-column:1 / -1;position:relative">'
        + '<div class="rc-headrow"><div class="rc-name">Total A/R Trend</div>'
        + '<div class="rc-desc">Weekly Total A/R (Current through 91+ Days), Brand vs. INK.</div>'
        + '<button class="rc-expand-btn" data-chart-id="chart-ar-weekly" data-chart-title="Total A/R Trend">⤢ Expand</button>'
        + '</div>'
        + '<div style="height:320px;margin-top:10px"><canvas id="chart-ar-weekly"></canvas></div>'
        + '</div>\n'
    )

def build_labor_chart_js(d):
    months = d['labor_total']['months']
    mon_labels = '[' + ','.join('"' + m + '"' for m,_,_ in months) + ']'
    plan_arr   = js_arr([round(p) for _,p,_ in months])
    act_arr    = '[' + ','.join(('null' if a is None else str(round(a))) for _,_,a in months) + ']'
    payroll    = d['payroll']
    ot_labels  = '[' + ','.join('"Wk ' + str(p['wk']) + '"' for p in payroll) + ']'
    ot_amt     = js_arr([round(p['ot_amount']) for p in payroll])
    ot_hrs     = js_arr([round(p['ot_hours'],1) for p in payroll])
    return (
        '(function(){const el=document.getElementById("laborMonthly");if(!el)return;'
        'new Chart(el,{type:"bar",data:{labels:' + mon_labels + ',datasets:['
        '{label:"Plan",data:' + plan_arr + ',backgroundColor:"#EDECED",borderRadius:3},'
        '{label:"Actual",data:' + act_arr + ',backgroundColor:"#43575E",borderRadius:3},'
        ']},options:{responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:true,position:"top"},tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmtK(c.raw)}}},'
        'scales:{x:{grid:{color:"#EDECED"}},y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}'
        '}});})();\n'
        '(function(){const el=document.getElementById("laborOT");if(!el)return;'
        'new Chart(el,{type:"line",data:{labels:' + ot_labels + ',datasets:['
        '{label:"OT $",data:' + ot_amt + ',borderColor:"#B4482F",backgroundColor:"transparent",borderWidth:2.5,pointRadius:3,tension:0.3,yAxisID:"y"},'
        '{label:"OT hrs",data:' + ot_hrs + ',borderColor:"#586D72",backgroundColor:"transparent",borderWidth:1.5,borderDash:[4,3],pointRadius:2,tension:0.3,yAxisID:"y1"},'
        ']},options:{responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:true,position:"top"},tooltip:{callbacks:{label:c=>c.dataset.label+": "+(c.dataset.yAxisID==="y"?fmtK(c.raw):c.raw+"h")}}},'
        'scales:{x:{grid:{color:"#EDECED"}},'
        'y:{position:"left",grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}},'
        'y1:{position:"right",grid:{drawOnChartArea:false},ticks:{callback:v=>v+"h"}}}'
        '}});})();\n'
    )

# ── OpEx panel ───────────────────────────────────────────────────────────────

def build_opex_panel(d):
    opex = d['opex']
    lines = d['opex_lines']
    keys = d['opex_keys']
    cats = d['opex_categories']

    body = rc_card_kpi("Total OpEx",
        (opex['plan'], opex['act'], opex['var']),
        (opex['ann_plan'], opex['ann_proj'], opex['ann_var']), favorable_if_below=True)

    if not cats:
        rows = [(lbl, plan, act, ann, ap, False) for lbl, act, plan, ann, ap in lines]
        body += rc_divider("OpEx by category")
        body += ('<div class="rc-card" style="grid-column:1 / -1">' + rc_hltable(rows, is_cost=True) + '</div>\n')
        return body

    groups = {}
    order = []
    for (lbl, act, plan, ann, ap), key in zip(lines, keys):
        meta = cats.get(key, {})
        cat_name = meta.get('category', 'Uncategorized')
        if cat_name not in groups:
            groups[cat_name] = []
            order.append(cat_name)
        groups[cat_name].append((lbl, act, plan, ann, ap, meta.get('highlight')))

    for cat_name in order:
        rows = [(lbl, plan, act, ann, ap, hl) for lbl, act, plan, ann, ap, hl in groups[cat_name]]
        sub_act  = sum(r[1] for r in groups[cat_name])
        sub_plan = sum(r[2] for r in groups[cat_name])
        sub_ann  = sum(r[3] for r in groups[cat_name])
        sub_ap   = sum(r[4] for r in groups[cat_name])
        body += rc_divider(cat_name)
        body += rc_card_kpi(cat_name,
            (sub_plan, sub_act, sub_act-sub_plan),
            (sub_ann, sub_ap, sub_ap-sub_ann), favorable_if_below=True,
            body_extra=rc_hltable(rows, is_cost=True))

    itemized_act      = sum(l[1] for l in lines)
    itemized_plan     = sum(l[2] for l in lines)
    itemized_ann_plan = sum(l[3] for l in lines)
    itemized_ann_proj = sum(l[4] for l in lines)
    other = (opex['act'] - itemized_act, opex['plan'] - itemized_plan,
             opex['ann_plan'] - itemized_ann_plan, opex['ann_proj'] - itemized_ann_proj)
    if any(abs(v) > 1 for v in other):
        body += rc_divider("Unallocated")
        body += ('<div class="rc-card" style="grid-column:1 / -1">'
                 + rc_hltable([("Other / unallocated OpEx", other[1], other[0], other[2], other[3], False)], is_cost=True)
                 + '</div>\n')
    return body

# ── Cash Flow panel ──────────────────────────────────────────────────────────

def build_ap_key_vendors_table(d):
    """Current open (unpaid) A/P - Key Vendors POs, by vendor -- terms,
    aging bucket, and the most-overdue PO on file for that vendor, from
    WHSL PMT Tracker's live per-PO detail."""
    kv = d['ap_key_vendors_open']
    terms = d.get('ap_key_vendors_terms') or {}
    vendors = kv['vendors']
    rows = ''
    for v in vendors:
        t = terms.get(v['vendor']) or '—'
        worst = v['worst_days']
        worst_lbl = (str(worst) + 'd overdue') if worst > 0 else 'not yet due'
        # Red marks what this vendor actually has to pay right now, so the
        # S&S exception (90+ only) reads straight off the table instead of
        # every vendor's 61-90 looking equally urgent.
        must_pay = kv_must_pay_buckets(v['vendor'])
        def aging_cell(bucket):
            style = 'color:var(--watch);font-weight:700' if bucket in must_pay else ''
            return '<td style="' + style + '">' + fk(v[bucket]) + '</td>'
        rows += (
            '<tr>'
            '<td style="text-align:left;font-family:var(--body)">' + html_escape(v['vendor']) + '</td>'
            '<td style="font-family:var(--body)">' + html_escape(t) + '</td>'
            + aging_cell('current')
            + aging_cell('d1_30')
            + aging_cell('d31_60')
            + aging_cell('d61_90')
            + aging_cell('d90_plus')
            + '<td style="font-weight:700">' + fk(v['total']) + '</td>'
            '<td style="font-family:var(--body);opacity:.75">' + worst_lbl + '</td>'
            '</tr>\n'
        )
    sums = {k: sum(v[k] for v in vendors) for k in ('current','d1_30','d31_60','d61_90','d90_plus','total')}
    rows += (
        '<tr style="border-top:1.5px solid var(--ink)">'
        '<td style="text-align:left;font-family:var(--body);font-weight:700;border-bottom:none">All Vendors</td>'
        '<td style="border-bottom:none"></td>'
        + ''.join('<td style="font-weight:700;border-bottom:none">' + fk(sums[k]) + '</td>'
                  for k in ('current','d1_30','d31_60','d61_90','d90_plus','total'))
        + '<td style="border-bottom:none"></td>'
        '</tr>\n'
    )
    return (
        '<div class="rc-hl">'
        '<table class="rc-hltable">'
        '<colgroup><col class="rc-hl-col-name"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat">'
        '<col class="rc-hl-col-stat"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat">'
        '<col class="rc-hl-col-stat"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat"></colgroup>'
        '<thead><tr><th>Vendor</th><th>Terms</th><th>Current</th><th>1-30</th><th>31-60</th>'
        '<th>61-90</th><th>90+</th><th>Total</th><th>Worst PO</th></tr></thead>\n'
        '<tbody>\n' + rows + '</tbody>\n'
        '</table></div>'
    )

def build_cashflow_panel(d):
    # Cash Flow tab rebuild, in progress -- Summary panel, Section #1 (Cash
    # In/Out by week, TY vs. LY), then the A/R trend. Everything else that
    # used to follow A/R (Cash Bridge, Bank Position, Cumulative Trend,
    # Credit Line, Paydown Feasibility, the A/P Key Vendors summary cards/
    # weekly chart, Cash In by Channel, Weekly Cash Collection Schedule,
    # Cash Out by Category) was removed per Jeremy (2026-09-04), keeping
    # only Open POs by Vendor.
    body = build_summary_panel(d)
    body += build_cf_ty_ly_section(d)
    body += build_ar_section(d)

    kv_open = d.get('ap_key_vendors_open')
    if kv_open:
        body += rc_divider("Open POs by Vendor")
        body += ('<div class="rc-card" style="grid-column:1 / -1">' + build_ap_key_vendors_table(d) + '</div>\n')

    return body

# ── Labor panel ──────────────────────────────────────────────────────────────

DEPT_LABELS = {
    'ink_design':'INK — Design','ink_ops':'INK — Operations','ink_prod':'INK — Production','ink_sales':'INK — Sales',
    'logistics':'Logistics','mkg':'Marketing','prodev':'Product Dev','flagship':'Flagship Store',
    'long_branch':'Long Branch','box_truck':'Box Truck Sales','whsl_sales':'Wholesale Sales','finance':'Finance',
}

def build_labor_panel(d):
    lt=d['labor_total']; li=d['labor_ink']; lb=d['labor_brand']

    body = rc_card_kpi("Total Labor",
        (lt['ytd_plan'],lt['ytd_act'],lt['ytd_var']), (lt['ann_plan'],lt['ann_proj'],lt['ann_var']),
        favorable_if_below=True,
        body_extra='<div class="rc-chart chart-wrap" style="height:200px;margin-top:14px"><canvas id="laborMonthly"></canvas></div>')
    body += rc_card_kpi("Brand Labor", (lb['ytd_plan'],lb['ytd_act'],lb['ytd_var']), (lb['ann_plan'],lb['ann_proj'],lb['ann_var']), favorable_if_below=True)
    body += rc_card_kpi("INK Labor",   (li['ytd_plan'],li['ytd_act'],li['ytd_var']), (li['ann_plan'],li['ann_proj'],li['ann_var']), favorable_if_below=True)

    pr = d['payroll_latest']
    if pr:
        body += rc_divider("Headcount — Pay Period Ending Week " + str(pr['wk']))
        pay_period_label = "Week " + str(pr['wk']) + (" — Pay Date " + pr['pay_date'] if pr['pay_date'] else "")
        body += (
            '<div class="rc-card" style="grid-column:1 / -1">'
            '<div class="rc-headrow"><div class="rc-name">Employment Mix</div></div>'
            '<div class="rc-group-label-fun">' + pay_period_label + '</div>'
            '<div class="rc-row-fun">' + build_labor_fun_stats(pr) + '</div>'
            '</div>\n'
        )
        dept_items = sorted(DEPT_LABELS.items(), key=lambda kv: -pr['depts'][kv[0]])
        dept_tiles = ''.join(dept_tile(lbl, str(pr['depts'][key])) for key, lbl in dept_items)
        body += (
            '<div class="rc-card" style="grid-column:1 / -1">'
            '<div class="dept-group-label">By Department</div>'
            '<div class="dept-grid">' + dept_tiles + '</div>'
            '</div>\n'
        )
        body += rc_divider("Overtime — Hours & Cost per Pay Period")
        ot_rows = ''.join(
            '<tr>'
            '<td style="text-align:left;font-family:var(--body)">' + (p['pay_date'] or ('Wk ' + str(p['wk']))) + '</td>'
            '<td>' + f"{p['ot_hours']:.1f}" + '</td>'
            '<td>' + fk(p['ot_amount']) + '</td>'
            '</tr>\n'
            for p in reversed(d['payroll'])
        )
        body += (
            '<div class="rc-card" style="grid-column:1 / -1">'
            '<div class="rc-chart chart-wrap" style="height:200px"><canvas id="laborOT"></canvas></div>'
            '<div class="rc-hl"><table class="rc-hltable">'
            '<colgroup><col class="rc-hl-col-name"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat"></colgroup>'
            '<thead><tr><th style="text-align:left">Pay Date</th><th>OT Hours</th><th>OT $</th></tr></thead>'
            '<tbody>' + ot_rows + '</tbody>'
            '</table></div>'
            '</div>\n'
        )
    return body

# ── Revenue panel ────────────────────────────────────────────────────────────

def build_revenue_panel(d):
    rev = d['rev']
    body = rc_card_kpi("Total Revenue",
        (rev['plan'], rev['act'], rev['var']), (rev['ann_plan'], rev['ann_proj'], rev['ann_var']),
        trend=d['rev_trend_total'])

    body += rc_divider("Revenue by Channel")
    cards = ''.join(ch_card(lbl, act, plan, ann, ap, tp) for lbl, act, plan, ann, ap, tp in d['rev_lines'])
    body += '<div class="channel-grid" style="grid-column:1 / -1">' + cards + '</div>\n'
    body += (
        '<div class="rc-note" style="grid-column:1 / -1">'
        'Trend Proj. adjusts each channel\'s full-year projection for its own seasonal shape, '
        'instead of assuming a flat pace to plan for the rest of the year. Curves are built from '
        'each channel\'s actual 2025 monthly split — e.g. Flagship Store\'s biggest month last year '
        'was December, not summer, the Mobile Store &amp; Tent pop-up did about a third of its year '
        'in October (the Box Truck warehouse sale), and Long Branch — which opened mid-2025, using '
        'real per-store daily sales from Jetty Hub since its books lumped it into Flagship — peaks '
        'in summer as expected of a shore-town shop.'
        '</div>\n'
    )

    oo = d['on_order']; per = d['on_order_periods']

    body += rc_divider("Open Orders — On Order, Not Yet Shipped or Invoiced")
    body += (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-headrow"><div class="rc-name">On Order — As of Week ' + str(d['week']) + '</div>'
        '<div class="rc-desc">Retail orders including discounts applied, not yet shipped or invoiced.</div></div>'
        '<div class="rc-boxrow"><div class="rc-box" style="flex:none;min-width:160px">'
        '<div class="rc-group-label">Total On Order</div>'
        '<div class="rc-stat-val" style="font-size:22px;margin-top:2px">' + fk(oo['total']) + '</div>'
        '</div></div>'
        '<div class="dept-group-label">By Category</div>'
        '<div class="dept-grid">' + ''.join(dept_tile(name, fk(amt)) for name, amt in d['on_order_cats']) + '</div>'
        '<div class="dept-group-label">By Half</div>'
        '<div class="dept-grid">'
        + dept_tile('H1 2026', fk(per['h1_2026']))
        + dept_tile('H2 2026', fk(per['h2_2026']))
        + dept_tile('H1 2027', fk(per['h1_2027']))
        + dept_tile('H2 2027', fk(per['h2_2027']))
        + '</div>'
        '</div>\n'
    )

    body += rc_divider("Wholesale — Plan vs. Booked, by Half (2026)")
    gap_rows = []
    for lbl, plan, act, order, proj in [
        ("H1 (Jan – Jun)", d['whsl_h1_plan'], d['whsl_h1_actual'], oo['h1'], d['whsl_h1_proj']),
        ("H2 (Jul – Dec)", d['whsl_h2_plan'], d['whsl_h2_actual'], oo['h2'], d['whsl_h2_proj']),
    ]:
        gap = plan - proj
        gap_cls = 'pos' if gap <= 0 else 'neg'
        gap_rows.append(
            '<tr>'
            '<td style="text-align:left;font-family:var(--body)">' + lbl + '</td>'
            '<td>' + fk(plan) + '</td>'
            '<td>' + fk(act) + '</td>'
            '<td>' + fk(order) + '</td>'
            '<td style="font-weight:600">' + fk(proj) + '</td>'
            '<td class="' + gap_cls + '" style="font-weight:700">' + (vk(-gap) if gap != 0 else '—') + '</td>'
            '</tr>\n'
        )
    body += (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-headrow"><div class="rc-name">Wholesale Gap Analysis</div>'
        '<div class="rc-desc">Projected = shipped/invoiced to date for that half, plus what\'s on order for it. '
        'Gap = plan minus projected — positive means still short of plan, negative means already ahead.</div></div>'
        '<div class="rc-hl"><table class="rc-hltable">'
        '<colgroup><col class="rc-hl-col-name"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat">'
        '<col class="rc-hl-col-stat"><col class="rc-hl-col-stat"><col class="rc-hl-col-stat"></colgroup>'
        '<thead><tr><th style="text-align:left">Half</th><th>Plan</th><th>Actual</th><th>On Order</th><th>Projected</th><th>Gap to Plan</th></tr></thead>'
        '<tbody>' + ''.join(gap_rows) + '</tbody>'
        '</table></div>'
        '</div>\n'
    )
    return body

# ── JRF panel ─────────────────────────────────────────────────────────────────
# Jetty Rock Foundation tracks its own budget in a separate, simpler workbook
# (monthly cadence, no weekly detail, no seasonality curves yet) -- see
# read_jrf(). This panel is only built (and only appears in the nav) when
# that file is present; see build_html.

def build_jrf_panel(jrf):
    as_of = str(jrf['as_of']).replace('As of', '').strip() if jrf['as_of'] else None
    body = (
        '<div class="rc-note" style="grid-column:1 / -1;padding-top:0">'
        'Jetty Rock Foundation tracks its own budget separately, on a monthly cadence'
        + (' — as of ' + as_of + '.' if as_of else '.')
        + '</div>\n'
    )

    body += rc_card_kpi("Total Revenue",
        (jrf['rev']['plan'], jrf['rev']['act'], jrf['rev']['var']),
        (jrf['rev']['ann_plan'], jrf['rev']['ann_proj'], jrf['rev']['ann_var']))
    body += rc_divider("Revenue by Category")
    cards = ''.join(ch_card(name, l['act'], l['plan'], l['ann_plan'], l['ann_proj']) for name, l in jrf['rev_cats'])
    body += '<div class="channel-grid" style="grid-column:1 / -1">' + cards + '</div>\n'

    body += rc_card_kpi("Donations (COGS)",
        (jrf['don']['plan'], jrf['don']['act'], jrf['don']['var']),
        (jrf['don']['ann_plan'], jrf['don']['ann_proj'], jrf['don']['ann_var']), favorable_if_below=True)
    body += rc_divider("Donations by Category")
    cards = ''.join(ch_card(name, l['act'], l['plan'], l['ann_plan'], l['ann_proj']) for name, l in jrf['don_cats'])
    body += '<div class="channel-grid" style="grid-column:1 / -1">' + cards + '</div>\n'

    body += rc_card_kpi("Total OpEx",
        (jrf['opex']['plan'], jrf['opex']['act'], jrf['opex']['var']),
        (jrf['opex']['ann_plan'], jrf['opex']['ann_proj'], jrf['opex']['ann_var']), favorable_if_below=True)
    body += rc_divider("OpEx by Category")
    cards = ''.join(ch_card(name, l['act'], l['plan'], l['ann_plan'], l['ann_proj']) for name, l in jrf['opex_cats'])
    body += '<div class="channel-grid" style="grid-column:1 / -1">' + cards + '</div>\n'

    body += rc_card_kpi("Net Income (Loss)",
        (jrf['net_income']['plan'], jrf['net_income']['act'], jrf['net_income']['var']),
        (jrf['net_income']['ann_plan'], jrf['net_income']['ann_proj'], jrf['net_income']['ann_var']))

    body += rc_divider("Revenue, Donations &amp; OpEx — Year over Year")
    body += (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-headrow"><div class="rc-name">Annual Totals, 2023–2025, vs. 2026 Year to Date</div>'
        '<div class="rc-desc">2026 is a partial year to date, not a full-year figure — expect it to look low.</div></div>'
        '<div class="rc-legend">'
        '<span><i class="swatch" style="background:#B0B4B8"></i>2023</span>'
        '<span><i class="swatch" style="background:#889096"></i>2024</span>'
        '<span><i class="swatch" style="background:#586D72"></i>2025</span>'
        '<span><i class="swatch" style="background:#43575E"></i>2026 YTD</span>'
        '</div>'
        '<div class="rc-chart chart-wrap" style="height:280px;margin-top:8px"><canvas id="jrfYoy"></canvas></div>'
        '</div>\n'
    )

    commit_rows = ''.join(
        '<tr><td style="text-align:left;font-family:var(--body)">' + name + '</td>'
        '<td style="padding-right:16px">' + fk(amt) + '</td>'
        '<td style="text-align:left;font-family:var(--body)">' + (bucket or '—') + '</td></tr>\n'
        for name, amt, bucket in jrf['commitments']
    )
    da = jrf['don_analysis']
    body += rc_divider("Outstanding Donation Commitments")
    body += (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-boxrow"><div class="rc-box"><div class="rc-group-label">Donated to Date</div><div class="rc-row">'
        + rc_stat('Donated', fk(da['donated_to_date']))
        + rc_stat('Projected', fk(da['projected']))
        + '</div></div>'
        '<div class="rc-box"><div class="rc-group-label">Commitments</div><div class="rc-row">'
        + rc_stat('Committed', fk(da['committed']))
        + rc_stat('Available to Donate', fk(da['available']))
        + '</div></div></div>'
        '<div class="rc-hl"><table class="rc-hltable">'
        '<colgroup><col class="rc-hl-col-name"><col class="rc-hl-col-stat"><col style="width:54.4%"></colgroup>'
        '<thead><tr><th style="text-align:left">Name</th><th style="padding-right:16px">Amount</th><th style="text-align:left">Bucket</th></tr></thead>'
        '<tbody>' + commit_rows
        + '<tr><td style="text-align:left;font-weight:700;font-family:var(--body)">Total</td>'
        + '<td style="font-weight:700">' + fk(jrf['commitments_total']) + '</td><td></td></tr></tbody>'
        '</table></div>'
        '</div>\n'
    )
    return body

def build_jrf_chart_js(jrf):
    years  = jrf['yoy_years']
    colors = ['#B0B4B8', '#889096', '#586D72', '#43575E']
    metrics = [
        ('Revenue',    jrf['yoy'].get('Total Income', [0,0,0,0])),
        ('Donations',  jrf['yoy'].get('Total Cost of Goods Sold', [0,0,0,0])),
        ('OpEx',       jrf['yoy'].get('Total Operating Expenses', [0,0,0,0])),
        ('Net Income', jrf['yoy'].get('Net Income', [0,0,0,0])),
    ]
    metric_labels = '[' + ','.join('"' + m + '"' for m, _ in metrics) + ']'
    datasets = ''.join(
        '{label:"' + yr + '",data:' + js_arr([round(vals[yi]) for _, vals in metrics]) + ',backgroundColor:"' + color + '",borderRadius:3},'
        for yi, (yr, color) in enumerate(zip(years, colors))
    )
    return (
        '(function(){const el=document.getElementById("jrfYoy");if(!el)return;'
        'new Chart(el,{type:"bar",data:{labels:' + metric_labels + ',datasets:[' + datasets + ']},'
        'options:{responsive:true,maintainAspectRatio:false,'
        'plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmtK(c.raw)}}},'
        'scales:{x:{grid:{color:"#EDECED"}},y:{grid:{color:"#EDECED"},ticks:{callback:v=>fmtK(v)}}}'
        '}});})();\n'
    )

# ── HTML assembly ─────────────────────────────────────────────────────────────

def sec_label(txt):
    return rc_divider(txt)

def instr_badge(cadence):
    """cadence: 'Weekly' | 'Bi-Weekly' | 'Monthly' -- colored pill matching
    the badge's own cadence so the eye can scan for "what's due this week"
    across sections."""
    cls = {'Weekly': 'weekly', 'Bi-Weekly': 'biweekly', 'Monthly': 'monthly'}[cadence]
    return '<span class="instr-badge instr-badge-' + cls + '">' + cadence + '</span>'

def instr_section(tab_label, cadence, sheet_note, items):
    lis = ''.join('<li>' + it + '</li>' for it in items)
    return (
        rc_divider(tab_label)
        + '<div class="rc-card" style="grid-column:1 / -1">'
        + '<div class="rc-headrow" style="display:flex;align-items:center;gap:10px">'
        + instr_badge(cadence)
        + ('<div class="rc-desc" style="margin:0">' + sheet_note + '</div>' if sheet_note else '')
        + '</div>'
        + '<ul class="instr-list">' + lis + '</ul>'
        + '</div>\n'
    )

def build_instructions_panel(d):
    """First-pass draft, inferred from which workbook tab feeds each
    dashboard section (see read_all()) -- cadences and exact steps are
    guesses about Jeremy's actual process and are meant to be corrected,
    not treated as settled fact."""
    intro = (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-headrow"><div class="rc-name">Keeping This Dashboard Current</div>'
        '<div class="rc-desc">This is a first-draft checklist, not a confirmed process — the cadence tags '
        '(Weekly / Bi-Weekly / Monthly) and specific steps below are guesses based on which tab in the budget '
        'workbook feeds each dashboard section. Correct anything that’s wrong or missing. One thing that is '
        'confirmed: once you save an edit to the workbook, the site checks for changes every 5 minutes and '
        'rebuilds automatically — there’s no separate publish step.</div></div>'
        '</div>\n'
    )
    sections = (
        instr_section('Revenue', 'Weekly',
            'Feeds the &ldquo;2026 Actual&rdquo; tab.',
            ['Enter this week’s actual revenue by channel (DTC, Wholesale, Screen Printing/INK, etc.).',
             'Reconcile the week’s total against Shopify/QuickBooks (or your source system) before saving.'])
        + instr_section('COGS &amp; Shipping', 'Weekly',
            'Also on the &ldquo;2026 Actual&rdquo; tab — Spring/Summer, Fall/Winter, and Special COGS, plus Brand/Ink Shipping.',
            ['Book this week’s COGS by season (Spring/Summer, Fall/Winter, Special).',
             'Enter Shipping actuals (Brand, Ink) for the week.'])
        + instr_section('Inventory', 'Weekly',
            'Feeds the &ldquo;Inventory&rdquo; and &ldquo;On Order&rdquo; tabs.',
            ['Update the on-hand COG/units snapshot for the week.',
             'Refresh sell-through % figures for the current season.',
             'Add or update any new POs on the &ldquo;On Order&rdquo; tab.'])
        + instr_section('Labor', 'Bi-Weekly',
            'Feeds the &ldquo;Labor&rdquo; and &ldquo;Payroll&rdquo; tabs, aligned to the pay cycle.',
            ['Enter headcount/hours by department for the completed pay period.',
             'Enter payroll cost actuals for the same period.'])
        + instr_section('OpEx', 'Monthly',
            'Feeds the OpEx line items on the &ldquo;2026 Actual&rdquo; tab (plus the optional &ldquo;Categories&rdquo; '
            'tab — see docs/opex-categories-setup.md).',
            ['Enter the month’s actual spend per OpEx line item as bills/invoices are booked.',
             'Add any new line item to the &ldquo;Categories&rdquo; tab so it doesn’t fall into Uncategorized.'])
        + instr_section('Cash Flow', 'Weekly',
            'Feeds the &ldquo;Cash Flow - Tracker&rdquo; tab.',
            ['Enter the week’s actual cash in/out.',
             'Confirm the forward schedule/forecast rows are still current.'])
    )
    return intro + sections

def build_html(d):
    WK  = d['week']
    rev = d['rev']; cogs = d['cogs']; opex = d['opex']; ni = d['net_income']

    # Revenue table (used by nothing now except cogs table below reuses the pattern)
    cogs_rows = [(lbl, plan, act, ann, ap, False) for lbl, act, plan, ann, ap in d['cogs_lines']]

    lw=d['inv_latest_wk']; pending=lw and lw<WK
    pend_note=" (wk " + str(lw) + " shown — wk " + str(WK) + " pending)" if pending else ""
    if lw:
        tc=d['cog_2026'].get(lw,0); tu=d['units_2026'].get(lw,0)
        s26_st=d['st_s26'].get(lw,0) or 0
        s26_py=d['st_s25'].get(lw,None)
        st_diff=round(s26_st-s26_py,1) if s26_py else None
        yoy_c=d['yoy_cog_2025']; yoy_u=d['yoy_units_2025']
        yoy_cog_str = ('+' + fk(tc-yoy_c) + ' vs prior year (+' + f'{(tc-yoy_c)/yoy_c*100:.1f}' + '%)') if yoy_c else ''
        yoy_u_str   = ('+' + f'{tu-yoy_u:,}' + ' units vs prior year (+' + f'{(tu-yoy_u)/yoy_u*100:.1f}' + '%)') if yoy_u else ''
        st_diff_str = ('S26 ' + f'{s26_st:.1f}' + '% — ' + f'{abs(st_diff):.1f}' + 'pts '
                       + ('behind' if st_diff<0 else 'ahead of') + ' S25 at wk ' + str(lw)) if st_diff is not None else ''
        prev_c = d['cog_2026'].get(d['inv_prev_wk'], 0)
        prev_u = d['units_2026'].get(d['inv_prev_wk'], 0)
        delta_c = tc - prev_c
        delta_c_str = ('+' if delta_c>=0 else '') + fk(delta_c) + ' vs prev wk'
    else:
        tc=tu=delta_c=0; yoy_cog_str=yoy_u_str=st_diff_str=delta_c_str=''

    legend_3yr = (
        '<div class="rc-legend">'
        '<span><i class="swatch y24" style="background:#B0B4B8"></i>2024</span>'
        '<span><i class="swatch y25" style="background:#586D72"></i>2025</span>'
        '<span><i class="swatch y26" style="background:#43575E"></i>2026</span>'
        '</div>'
    )

    qtr_wk_pct = WK / 52 * 100
    milestones = [13, 26, 39, 52]
    qtr_names = ['Q1','Q2','Q3','Q4']
    cur_q = min(3, (WK - 1) // 13)
    qtr_labels_html = ''.join(
        '<span class="' + ('done' if i < cur_q else ('current' if i == cur_q else 'upcoming')) + '">' + qtr_names[i] + '</span>'
        for i in range(4)
    )
    milestones_html = ''.join(
        '<div class="qtr-milestone' + (' passed' if WK > mw else '') + '" style="left:' + str(pct) + '%"></div>'
        for mw, pct in zip(milestones, [25, 50, 75, 100])
    )

    CSS = FONTS_CSS + RC_CSS + EXTRA_CSS

    topbar = (
        '<header class="topbar">\n'
        '  <div class="left-group">\n'
        '    <div class="brand">\n'
        '      <img src="data:image/png;base64,' + LOGO_B64 + '">\n'
        '      <div class="divider"></div>\n'
        '      <div><div class="title-main">Strategic Dashboard</div>'
        '<div class="title-sub">2026 · YTD Cumulative</div></div>\n'
        '    </div>\n'
        '    <div class="divider"></div>\n'
        '    <div class="week-pill"><span class="lbl">Thru</span><span class="val">Week ' + str(WK) + '</span><span class="of">of 52</span></div>\n'
        '  </div>\n'
        '  <div class="qtr-timeline">\n'
        '    <div class="qtr-track">\n'
        '      <div class="qtr-fill" style="width:' + f'{qtr_wk_pct:.1f}' + '%"></div>\n'
        + milestones_html +
        '      <div class="qtr-now" style="left:' + f'{qtr_wk_pct:.1f}' + '%" title="Week ' + str(WK) + '"></div>\n'
        '    </div>\n'
        '    <div class="qtr-labels">' + qtr_labels_html + '</div>\n'
        '  </div>\n'
        '</header>\n'
    )

    tabs = (
        '<nav class="tabs">\n'
        '  <div class="tab active" data-panel="summary">Summary</div>\n'
        '  <div class="tab" data-panel="revenue">Revenue</div>\n'
        '  <div class="tab" data-panel="cogs">COGS &amp; Shipping</div>\n'
        '  <div class="tab" data-panel="inventory">Inventory</div>\n'
        '  <div class="tab" data-panel="labor">Labor</div>\n'
        '  <div class="tab" data-panel="opex">OpEx</div>\n'
        '  <div class="tab" data-panel="cashflow">Cash Flow</div>\n'
        + ('  <div class="tab" data-panel="jrf">JRF</div>\n' if d.get('jrf') else '') +
        '  <div class="tab tab-instructions" data-panel="instructions">Instructions</div>\n'
        '</nav>\n'
    )

    trend_chart_desc = (
        "2025 Actual (gray) is last year's cumulative pace for reference. 2026 Trend (dashed) "
        "continues the solid 2026 Actual line from this week forward."
    )
    summary_panel = (
        rc_card_kpi("Total Revenue", (rev['plan'],rev['act'],rev['var']), (rev['ann_plan'],rev['ann_proj'],rev['ann_var']),
            trend=d['rev_trend_total'])
        + rc_trend_chart('chart-trend-revenue', '52-Week Trend — Total Revenue', trend_chart_desc)
        + rc_card_kpi("COGS + Labor + Shipping", (cogs['plan'],cogs['act'],cogs['var']), (cogs['ann_plan'],cogs['ann_proj'],cogs['ann_var']),
            favorable_if_below=True)
        + rc_trend_chart('chart-trend-cogs', '52-Week Trend — COGS + Labor + Shipping', trend_chart_desc)
        + rc_card_kpi("Total OpEx", (opex['plan'],opex['act'],opex['var']), (opex['ann_plan'],opex['ann_proj'],opex['ann_var']),
            favorable_if_below=True)
        + rc_trend_chart('chart-trend-opex', '52-Week Trend — Total OpEx', trend_chart_desc)
        + rc_card_kpi("Net Income", (ni['plan'],ni['act'],ni['var']), (ni['ann_plan'],ni['ann_proj'],ni['ann_var']),
            trend=d['net_income_trend'],
            desc="Trend uses Revenue's seasonality curve; COGS/OpEx use their plan-paced Full Year Projection.")
        + rc_trend_chart('chart-trend-net_income', '52-Week Trend — Net Income', trend_chart_desc)
    )

    cogs_panel = (
        rc_card_kpi("Total COGS + Labor + Shipping",
            (cogs['plan'],cogs['act'],cogs['var']), (cogs['ann_plan'],cogs['ann_proj'],cogs['ann_var']),
            favorable_if_below=True, body_extra=rc_hltable(cogs_rows, is_cost=True))
    )

    avg_unit_cost = (tc / tu) if tu else 0
    note_parts = []
    if delta_c_str: note_parts.append('COG ' + delta_c_str)
    if yoy_cog_str: note_parts.append('COG ' + yoy_cog_str)
    if yoy_u_str:   note_parts.append(yoy_u_str)
    inv_note = ('. '.join(note_parts) + '.') if note_parts else 'No comparison data available.'

    inventory_panel = (
        '<div class="rc-card" style="grid-column:1 / -1">'
        '<div class="rc-headrow"><div class="rc-name">Inventory Snapshot — Week ' + str(lw or WK) + pend_note + '</div>'
        '<div class="rc-desc">' + inv_note + '</div></div>'
        '<div class="rc-boxrow"><div class="rc-box"><div class="rc-group-label">On Hand</div><div class="rc-row">'
        + rc_stat('Total COG', fk(tc) if tc else '—')
        + rc_stat('Units', format(tu, ',') if tu else '—')
        + rc_stat('Avg Unit Cost', '$' + f'{avg_unit_cost:,.2f}' if tu else '—')
        + '</div></div></div>'
        '</div>\n'
        + rc_divider("Total Inventory COG — Week by Week")
        + '<div class="rc-card" style="grid-column:1 / -1">'
        + '<div class="rc-headrow"><div class="rc-name">' + (yoy_cog_str or ' ') + '</div></div>'
        + legend_3yr
        + '<div class="rc-chart chart-wrap" style="height:260px;margin-top:8px"><canvas id="invHistCOG"></canvas></div>'
        + '</div>\n'
        + rc_divider("Total Units on Hand — Week by Week")
        + '<div class="rc-card" style="grid-column:1 / -1">'
        + '<div class="rc-headrow"><div class="rc-name">' + (yoy_u_str or ' ') + '</div></div>'
        + legend_3yr
        + '<div class="rc-chart chart-wrap" style="height:260px;margin-top:8px"><canvas id="invTotUnits"></canvas></div>'
        + '</div>\n'
        + rc_divider("Spring / Summer Sell-Through % — by Week")
        + '<div class="rc-card" style="grid-column:1 / -1">'
        + '<div class="rc-headrow"><div class="rc-name">' + (st_diff_str or ' ') + '</div></div>'
        + '<div class="rc-legend"><span><i class="swatch" style="background:#B0B4B8"></i>S24</span><span><i class="swatch" style="background:#586D72"></i>S25</span><span><i class="swatch" style="background:#43575E"></i>S26</span></div>'
        + '<div class="rc-chart chart-wrap" style="height:240px;margin-top:8px"><canvas id="invSSSellThru"></canvas></div>'
        + '</div>\n'
        + rc_divider("Fall / Winter Sell-Through % — by Week")
        + '<div class="rc-card" style="grid-column:1 / -1">'
        + '<div class="rc-legend"><span><i class="swatch" style="background:#B0B4B8"></i>F24</span><span><i class="swatch" style="background:#586D72"></i>F25</span><span><i class="swatch" style="background:#43575E"></i>F26</span></div>'
        + '<div class="rc-chart chart-wrap" style="height:240px;margin-top:8px"><canvas id="invFWSellThru"></canvas></div>'
        + '</div>\n'
    )

    def panel(pid, content):
        return '<div class="tab-panel' + (' active' if pid=='summary' else '') + '" id="panel-' + pid + '">\n<div class="rc-grid">\n' + content + '</div>\n</div>\n\n'

    return (
        '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
        '<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n'
        '<title>Jetty Strategic Dashboard 2026 — Week ' + str(WK) + '</title>\n'
        '<style>' + CSS + '</style>\n</head>\n<body>\n'
        '<div class="page">\n\n'
        + topbar + tabs +
        '\n<main class="content">\n'
        + panel('summary', summary_panel)
        + panel('revenue', build_revenue_panel(d))
        + panel('cogs', cogs_panel)
        + panel('inventory', inventory_panel)
        + panel('labor', build_labor_panel(d))
        + panel('opex', build_opex_panel(d))
        + panel('cashflow', build_cashflow_panel(d))
        + (panel('jrf', build_jrf_panel(d['jrf'])) if d.get('jrf') else '')
        + panel('instructions', build_instructions_panel(d))
        + '</main>\n'
        '</div>\n'
        + CHART_MODAL_HTML +
        '<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>\n'
        '<script>\n'
        + build_chart_js(d)
        + '</script>\n'
        '<script>\n'
        + TAB_JS +
        '</script>\n'
        '<script>\n'
        + CHART_MODAL_JS +
        '</script>\n</body>\n</html>'
    )

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Reading data...")
    d = read_all()
    print("Week " + str(d['week']) + " data loaded")
    d['jrf'] = read_jrf()
    print("JRF data loaded" if d['jrf'] else "JRF file not found — skipping JRF tab")
    print("Building dashboard...")
    html = build_html(d)
    os.makedirs("output", exist_ok=True)
    with open("output/index.html", "w") as f:
        f.write(html)
    size = os.path.getsize("output/index.html") / 1024
    print("Done — output/index.html written (" + str(round(size,1)) + " KB)")

    source_modified = os.environ.get("SOURCE_MODIFIED_TIME", "")
    jrf_modified = os.environ.get("JRF_MODIFIED_TIME", "")
    with open("output/meta.json", "w") as f:
        json.dump({"source_modified": source_modified, "jrf_modified": jrf_modified}, f)
    print("Recorded source_modified=" + source_modified + ", jrf_modified=" + jrf_modified + " in output/meta.json")

if __name__ == "__main__":
    main()
